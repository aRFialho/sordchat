from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Optional
import csv
import io
import json
import os
import re
import uuid

from fastapi import Depends, FastAPI, File, HTTPException, Query, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy import BigInteger, Boolean, Column, DateTime, ForeignKey, Integer, LargeBinary, String, Text, create_engine, or_, text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship, sessionmaker
import uvicorn


SECRET_KEY = os.getenv("SECRET_KEY", "voltcorp_secret_key_super_secure_2024")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "1440"))
IS_PRODUCTION = os.getenv("ENVIRONMENT", "development").lower() == "production"
APP_VERSION = os.getenv("APP_VERSION") or os.getenv("RENDER_GIT_COMMIT") or "local"
APP_BUILD_TIME = os.getenv("APP_BUILD_TIME")
DEFAULT_FRONTEND_ORIGINS = [
    "https://voltcorp-web.onrender.com",
]
DEFAULT_DEPARTMENTS = ["TI", "Suporte", "Comercial", "Financeiro", "Operacao", "Produto"]
DEFAULT_COMPANY_ID = "00000000-0000-0000-0000-000000000001"
DEFAULT_COMPANY_NAME = "Empresa Padrao"
USER_LEVELS = {"usuario", "coordenador", "master", "padrao"}
COORDINATOR_LEVELS = {"coordenador", "master"}
PLATFORM_ROLES = {"master_admin", "company_admin", "coordinator", "user"}
TENANT_ADMIN_ROLES = {"company_admin", "master_admin"}
ACTIVE_STATUS = "active"
INACTIVE_STATUS = "inactive"
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
ALLOWED_UPLOAD_EXTENSIONS = {
    ".csv",
    ".doc",
    ".docx",
    ".gif",
    ".jpeg",
    ".jpg",
    ".pdf",
    ".png",
    ".txt",
    ".webp",
    ".xls",
    ".xlsx",
    ".zip",
}


def normalize_database_url(database_url: str) -> str:
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    if database_url.startswith("postgresql://"):
        database_url = database_url.replace("postgresql://", "postgresql+psycopg://", 1)
    return database_url


DATABASE_URL = normalize_database_url(os.getenv("DATABASE_URL", "sqlite:///./voltcorp.db"))
engine_options = {"pool_pre_ping": True}
if DATABASE_URL.startswith("sqlite"):
    engine_options["connect_args"] = {"check_same_thread": False}

engine = create_engine(DATABASE_URL, **engine_options)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()


class User(Base):
    __tablename__ = "users"

    id = Column(Integer, primary_key=True, index=True)
    uuid = Column(String(36), default=lambda: str(uuid.uuid4()), unique=True, nullable=True, index=True)
    username = Column(String(80), unique=True, index=True, nullable=False)
    email = Column(String(255), unique=True, index=True, nullable=False)
    full_name = Column(String(255), nullable=False)
    hashed_password = Column(String(255), nullable=False)
    phone = Column(String(40), nullable=True)
    is_platform_admin = Column(Boolean, default=False, nullable=False)
    must_change_password = Column(Boolean, default=False, nullable=False)
    status = Column(String(40), default=ACTIVE_STATUS, nullable=False)
    access_level = Column(String(40), default="usuario", nullable=False)
    department = Column(String(100), nullable=True)
    phone_extension = Column(String(40), nullable=True)
    birthday = Column(String(10), nullable=True)
    role_title = Column(String(100), nullable=True)
    is_active = Column(Boolean, default=True, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(DateTime, nullable=True)

    sent_messages = relationship("Message", foreign_keys="Message.sender_id", back_populates="sender")
    received_messages = relationship("Message", foreign_keys="Message.receiver_id", back_populates="receiver")
    uploaded_files = relationship("FileUpload", back_populates="uploader")


class Company(Base):
    __tablename__ = "companies"

    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    name = Column(String(255), nullable=False, index=True)
    cnpj = Column(String(32), nullable=True, unique=True, index=True)
    responsible_name = Column(String(255), nullable=True)
    phone_primary = Column(String(40), nullable=True)
    phone_secondary = Column(String(40), nullable=True)
    status = Column(String(40), default=ACTIVE_STATUS, nullable=False, index=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(DateTime, nullable=True)


class Department(Base):
    __tablename__ = "departments"

    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    company_id = Column(String(36), ForeignKey("companies.id"), nullable=False, index=True)
    name = Column(String(120), nullable=False, index=True)
    description = Column(Text, nullable=True)
    status = Column(String(40), default=ACTIVE_STATUS, nullable=False, index=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)


class CompanyUser(Base):
    __tablename__ = "company_users"

    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    company_id = Column(String(36), ForeignKey("companies.id"), nullable=False, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False, index=True)
    department_id = Column(String(36), ForeignKey("departments.id"), nullable=True, index=True)
    role = Column(String(40), default="user", nullable=False, index=True)
    status = Column(String(40), default=ACTIVE_STATUS, nullable=False, index=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)


class AuditLog(Base):
    __tablename__ = "audit_logs"

    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    company_id = Column(String(36), ForeignKey("companies.id"), nullable=True, index=True)
    actor_user_id = Column(Integer, ForeignKey("users.id"), nullable=True, index=True)
    action = Column(String(100), nullable=False, index=True)
    entity_type = Column(String(100), nullable=False, index=True)
    entity_id = Column(String(80), nullable=True, index=True)
    metadata_json = Column("metadata", Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)


class Ticket(Base):
    __tablename__ = "tickets"

    id = Column(Integer, primary_key=True, index=True)
    company_id = Column(String(36), ForeignKey("companies.id"), default=DEFAULT_COMPANY_ID, nullable=True, index=True)
    title = Column(String(200), nullable=False, index=True)
    description = Column(Text, nullable=False)
    priority = Column(String(40), default="Media", nullable=False)
    status = Column(String(40), default="Aberto", nullable=False)
    department = Column(String(100), nullable=True, index=True)
    channel = Column(String(80), default="Web", nullable=False)
    created_by_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    assigned_to_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    attachment_file_id = Column(Integer, ForeignKey("file_uploads.id"), nullable=True)
    rating_score = Column(Integer, nullable=True)
    rating_comment = Column(Text, nullable=True)
    rated_at = Column(DateTime, nullable=True)
    first_response_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(DateTime, nullable=True)
    closed_at = Column(DateTime, nullable=True)


class ChatGroup(Base):
    __tablename__ = "chat_groups"

    id = Column(Integer, primary_key=True, index=True)
    company_id = Column(String(36), ForeignKey("companies.id"), default=DEFAULT_COMPANY_ID, nullable=True, index=True)
    name = Column(String(120), nullable=False, index=True)
    description = Column(Text, nullable=True)
    department = Column(String(100), nullable=True, index=True)
    created_by_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    is_active = Column(Boolean, default=True, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)


class TaskItem(Base):
    __tablename__ = "tasks"

    id = Column(Integer, primary_key=True, index=True)
    company_id = Column(String(36), ForeignKey("companies.id"), default=DEFAULT_COMPANY_ID, nullable=True, index=True)
    title = Column(String(200), nullable=False, index=True)
    description = Column(Text, nullable=True)
    priority = Column(String(40), default="medium", nullable=False)
    category = Column(String(100), default="Operacao", nullable=False)
    status = Column(String(40), default="backlog", nullable=False, index=True)
    due_date = Column(String(10), nullable=True)
    created_by_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    assigned_to_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(DateTime, nullable=True)


class Message(Base):
    __tablename__ = "messages"

    id = Column(Integer, primary_key=True, index=True)
    company_id = Column(String(36), ForeignKey("companies.id"), default=DEFAULT_COMPANY_ID, nullable=True, index=True)
    content = Column(Text, nullable=False)
    sender_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    receiver_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    message_type = Column(String(40), default="text", nullable=False)
    file_path = Column(String(500), nullable=True)
    timestamp = Column(DateTime, default=datetime.utcnow, nullable=False)
    is_read = Column(Boolean, default=False, nullable=False)

    sender = relationship("User", foreign_keys=[sender_id], back_populates="sent_messages")
    receiver = relationship("User", foreign_keys=[receiver_id], back_populates="received_messages")


class TicketMessage(Base):
    __tablename__ = "ticket_messages"

    id = Column(Integer, primary_key=True, index=True)
    company_id = Column(String(36), ForeignKey("companies.id"), default=DEFAULT_COMPANY_ID, nullable=True, index=True)
    ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False, index=True)
    sender_id = Column(Integer, ForeignKey("users.id"), nullable=False, index=True)
    content = Column(Text, default="", nullable=False)
    file_id = Column(Integer, ForeignKey("file_uploads.id"), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)


class FileUpload(Base):
    __tablename__ = "file_uploads"

    id = Column(Integer, primary_key=True, index=True)
    company_id = Column(String(36), ForeignKey("companies.id"), default=DEFAULT_COMPANY_ID, nullable=True, index=True)
    filename = Column(String(255), nullable=False)
    file_path = Column(String(500), nullable=False)
    file_size = Column(Integer, nullable=False)
    content_type = Column(String(120), nullable=False)
    binary_data = Column(LargeBinary, nullable=True)
    uploaded_by = Column(Integer, ForeignKey("users.id"), nullable=False)
    upload_date = Column(DateTime, default=datetime.utcnow, nullable=False)

    uploader = relationship("User", back_populates="uploaded_files")


class DesktopRelease(Base):
    __tablename__ = "desktop_releases"

    id = Column(Integer, primary_key=True, index=True)
    version = Column(String(80), nullable=False, index=True)
    platform = Column(String(40), default="windows", nullable=False, index=True)
    filename = Column(String(255), nullable=False)
    content_type = Column(String(120), default="application/octet-stream", nullable=False)
    file_size = Column(BigInteger, nullable=False)
    sha256 = Column(String(64), nullable=False)
    binary_data = Column(LargeBinary, nullable=True)
    storage_mode = Column(String(40), default="inline", nullable=False)
    is_active = Column(Boolean, default=True, nullable=False, index=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)


class DesktopReleaseChunk(Base):
    __tablename__ = "desktop_release_chunks"

    release_id = Column(Integer, ForeignKey("desktop_releases.id", ondelete="CASCADE"), primary_key=True)
    chunk_index = Column(Integer, primary_key=True)
    data = Column(LargeBinary, nullable=False)


def sqlite_table_columns(conn, table_name: str) -> set[str]:
    rows = conn.exec_driver_sql(f"PRAGMA table_info({table_name})").fetchall()
    return {row[1] for row in rows}


def sqlite_add_column_if_missing(conn, table_name: str, column_name: str, ddl: str):
    if column_name not in sqlite_table_columns(conn, table_name):
        conn.exec_driver_sql(f"ALTER TABLE {table_name} ADD COLUMN {ddl}")


def ensure_sqlite_multi_tenant_schema():
    with engine.begin() as conn:
        user_columns = sqlite_table_columns(conn, "users")
        if "uuid" not in user_columns:
            conn.exec_driver_sql("ALTER TABLE users ADD COLUMN uuid VARCHAR(36)")
            rows = conn.exec_driver_sql("SELECT id FROM users WHERE uuid IS NULL").fetchall()
            for row in rows:
                conn.execute(text("UPDATE users SET uuid = :uuid WHERE id = :id"), {"uuid": str(uuid.uuid4()), "id": row[0]})
        if "phone" not in user_columns:
            conn.exec_driver_sql("ALTER TABLE users ADD COLUMN phone VARCHAR(40)")
        if "is_platform_admin" not in user_columns:
            conn.exec_driver_sql("ALTER TABLE users ADD COLUMN is_platform_admin BOOLEAN DEFAULT 0 NOT NULL")
        if "must_change_password" not in user_columns:
            conn.exec_driver_sql("ALTER TABLE users ADD COLUMN must_change_password BOOLEAN DEFAULT 0 NOT NULL")
        if "status" not in user_columns:
            conn.exec_driver_sql("ALTER TABLE users ADD COLUMN status VARCHAR(40) DEFAULT 'active' NOT NULL")

        for table_name in ["tickets", "chat_groups", "tasks", "messages", "ticket_messages", "file_uploads"]:
            if table_name in {"ticket_messages", "tasks", "chat_groups", "tickets", "messages", "file_uploads"}:
                sqlite_add_column_if_missing(conn, table_name, "company_id", "company_id VARCHAR(36)")
                conn.execute(
                    text(f"UPDATE {table_name} SET company_id = :company_id WHERE company_id IS NULL"),
                    {"company_id": DEFAULT_COMPANY_ID},
                )


if DATABASE_URL.startswith("sqlite"):
    Base.metadata.create_all(bind=engine)
    ensure_sqlite_multi_tenant_schema()


def get_cors_origins():
    configured = os.getenv("FRONTEND_ORIGINS", "")
    origins = list(DEFAULT_FRONTEND_ORIGINS)
    if configured:
        origins.extend(origin.strip() for origin in configured.split(",") if origin.strip())
    if not IS_PRODUCTION:
        origins.extend(["http://127.0.0.1:3000", "http://localhost:3000"])
    return sorted(set(origins))


app = FastAPI(title="Volt Corp API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=get_cors_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def enforce_required_password_change(request, call_next):
    allowed_paths = {
        "/",
        "/health",
        "/version",
        "/auth/login",
        "/auth/logout",
        "/auth/me",
        "/auth/change-password",
    }
    path = request.url.path
    if path in allowed_paths or path.startswith("/downloads/"):
        return await call_next(request)

    authorization = request.headers.get("authorization") or ""
    if authorization.lower().startswith("bearer "):
        token = authorization.split(" ", 1)[1].strip()
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            user_id = payload.get("sub")
            if user_id:
                with SessionLocal() as db:
                    user = db.query(User).filter(User.id == int(user_id)).first()
                    if user and user.must_change_password:
                        return JSONResponse(
                            status_code=403,
                            content={
                                "detail": "Troca de senha obrigatoria no primeiro acesso.",
                                "must_change_password": True,
                            },
                        )
        except JWTError:
            pass

    return await call_next(request)


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=15))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token invalido")
    except JWTError:
        raise HTTPException(status_code=401, detail="Token invalido")

    with SessionLocal() as db:
        user = db.query(User).filter(User.id == int(user_id)).first()
        if user is None:
            raise HTTPException(status_code=401, detail="Usuario nao encontrado")
        db.expunge(user)
        return user


def normalize_access_level(access_level: Optional[str]) -> str:
    value = (access_level or "usuario").strip().lower()
    if value == "padrao":
        return "usuario"
    return value if value in USER_LEVELS else "usuario"


def normalize_status(status: Optional[str]) -> str:
    value = (status or ACTIVE_STATUS).strip().lower()
    if value in {"ativo", "active", "enabled", "habilitado"}:
        return ACTIVE_STATUS
    if value in {"inativo", "inactive", "disabled", "desabilitado"}:
        return INACTIVE_STATUS
    return value or ACTIVE_STATUS


def normalize_platform_role(role: Optional[str]) -> str:
    value = (role or "user").strip().lower()
    aliases = {
        "master": "master_admin",
        "admin_master": "master_admin",
        "admin": "company_admin",
        "administrador": "company_admin",
        "coordenador": "coordinator",
        "usuario": "user",
    }
    value = aliases.get(value, value)
    return value if value in PLATFORM_ROLES else "user"


def legacy_access_for_role(role: str) -> str:
    role = normalize_platform_role(role)
    if role == "coordinator":
        return "coordenador"
    return "usuario"


def is_admin(user: User) -> bool:
    return bool(getattr(user, "is_platform_admin", False)) or normalize_access_level(user.access_level) == "master"


def is_coordinator(user: User) -> bool:
    return normalize_access_level(user.access_level) in COORDINATOR_LEVELS


def ensure_admin(user: User):
    if not is_admin(user):
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador.")


def ensure_coordinator(user: User):
    if not is_coordinator(user):
        raise HTTPException(status_code=403, detail="Acesso restrito a coordenadores.")


def serialize_company(company: Company) -> dict:
    return {
        "id": company.id,
        "name": company.name,
        "cnpj": company.cnpj,
        "responsible_name": company.responsible_name,
        "phone_primary": company.phone_primary,
        "phone_secondary": company.phone_secondary,
        "status": company.status,
        "created_at": company.created_at.isoformat() if company.created_at else None,
        "updated_at": company.updated_at.isoformat() if company.updated_at else None,
    }


def serialize_department(department: Department) -> dict:
    return {
        "id": department.id,
        "company_id": department.company_id,
        "name": department.name,
        "description": department.description,
        "status": department.status,
        "created_at": department.created_at.isoformat() if department.created_at else None,
    }


def serialize_company_user(company_user: CompanyUser, db) -> dict:
    company = db.query(Company).filter(Company.id == company_user.company_id).first()
    user = db.query(User).filter(User.id == company_user.user_id).first()
    department = db.query(Department).filter(Department.id == company_user.department_id).first() if company_user.department_id else None
    return {
        "id": company_user.id,
        "company_id": company_user.company_id,
        "company_name": company.name if company else None,
        "user_id": company_user.user_id,
        "user_name": user.full_name if user else None,
        "user_email": user.email if user else None,
        "department_id": company_user.department_id,
        "department_name": department.name if department else None,
        "role": company_user.role,
        "status": company_user.status,
        "created_at": company_user.created_at.isoformat() if company_user.created_at else None,
    }


def serialize_audit_log(log: AuditLog) -> dict:
    metadata = {}
    if log.metadata_json:
        try:
            metadata = json.loads(log.metadata_json)
        except json.JSONDecodeError:
            metadata = {"raw": log.metadata_json}
    return {
        "id": log.id,
        "company_id": log.company_id,
        "actor_user_id": log.actor_user_id,
        "action": log.action,
        "entity_type": log.entity_type,
        "entity_id": log.entity_id,
        "metadata": metadata,
        "created_at": log.created_at.isoformat() if log.created_at else None,
    }


def serialize_user(user: User, db=None) -> dict:
    memberships = []
    if db is not None:
        memberships = [
            serialize_company_user(item, db)
            for item in db.query(CompanyUser)
            .filter(CompanyUser.user_id == user.id, CompanyUser.status == ACTIVE_STATUS)
            .order_by(CompanyUser.created_at.asc())
            .all()
        ]

    return {
        "id": user.id,
        "uuid": user.uuid,
        "name": user.full_name,
        "username": user.username,
        "email": user.email,
        "full_name": user.full_name,
        "phone": getattr(user, "phone", None) or user.phone_extension,
        "is_platform_admin": bool(getattr(user, "is_platform_admin", False)),
        "must_change_password": bool(getattr(user, "must_change_password", False)),
        "status": getattr(user, "status", ACTIVE_STATUS) or (ACTIVE_STATUS if user.is_active else INACTIVE_STATUS),
        "access_level": normalize_access_level(user.access_level),
        "department": user.department,
        "phone_extension": user.phone_extension,
        "birthday": user.birthday,
        "role_title": user.role_title,
        "is_active": user.is_active,
        "companies": memberships,
        "company_id": memberships[0]["company_id"] if memberships else None,
        "company_role": memberships[0]["role"] if memberships else ("master_admin" if is_admin(user) else None),
        "company_name": memberships[0]["company_name"] if memberships else None,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "updated_at": user.updated_at.isoformat() if user.updated_at else None,
    }


def log_audit(db, actor_user_id: Optional[int], action: str, entity_type: str, entity_id: Optional[str], company_id: Optional[str] = None, metadata: Optional[dict] = None):
    db.add(
        AuditLog(
            company_id=company_id,
            actor_user_id=actor_user_id,
            action=action,
            entity_type=entity_type,
            entity_id=str(entity_id) if entity_id is not None else None,
            metadata_json=json.dumps(metadata or {}, ensure_ascii=False),
        )
    )


def get_or_create_department(db, company_id: str, name: str, description: Optional[str] = None) -> Department:
    clean_name = str(name or "").strip() or "Geral"
    department = (
        db.query(Department)
        .filter(Department.company_id == company_id, Department.name.ilike(clean_name))
        .first()
    )
    if department:
        return department

    department = Department(
        id=str(uuid.uuid4()),
        company_id=company_id,
        name=clean_name,
        description=description,
        status=ACTIVE_STATUS,
    )
    db.add(department)
    db.flush()
    return department


def ensure_default_tenant_records():
    with SessionLocal() as db:
        company = db.query(Company).filter(Company.id == DEFAULT_COMPANY_ID).first()
        if not company:
            company = Company(
                id=DEFAULT_COMPANY_ID,
                name=DEFAULT_COMPANY_NAME,
                responsible_name="Admin Master",
                status=ACTIVE_STATUS,
            )
            db.add(company)
            db.flush()

        department_by_name = {}
        for name in sorted(set(DEFAULT_DEPARTMENTS + ["Administracao"])):
            department = get_or_create_department(db, DEFAULT_COMPANY_ID, name)
            department_by_name[name.lower()] = department

        for user in db.query(User).all():
            if not user.uuid:
                user.uuid = str(uuid.uuid4())
            if normalize_access_level(user.access_level) == "master" and not user.is_platform_admin:
                user.is_platform_admin = True
            if not getattr(user, "status", None):
                user.status = ACTIVE_STATUS if user.is_active else INACTIVE_STATUS
            department = department_by_name.get((user.department or "Operacao").lower()) or get_or_create_department(
                db, DEFAULT_COMPANY_ID, user.department or "Operacao"
            )
            membership = (
                db.query(CompanyUser)
                .filter(CompanyUser.company_id == DEFAULT_COMPANY_ID, CompanyUser.user_id == user.id)
                .first()
            )
            if not membership:
                role = "master_admin" if user.is_platform_admin else ("coordinator" if normalize_access_level(user.access_level) == "coordenador" else "user")
                db.add(
                    CompanyUser(
                        id=str(uuid.uuid4()),
                        company_id=DEFAULT_COMPANY_ID,
                        user_id=user.id,
                        department_id=department.id,
                        role=role,
                        status=ACTIVE_STATUS if user.is_active else INACTIVE_STATUS,
                    )
                )
        db.commit()


def active_company_memberships(db, user: User) -> list[CompanyUser]:
    return (
        db.query(CompanyUser)
        .join(Company, Company.id == CompanyUser.company_id)
        .filter(
            CompanyUser.user_id == user.id,
            CompanyUser.status == ACTIVE_STATUS,
            Company.status == ACTIVE_STATUS,
        )
        .order_by(CompanyUser.created_at.asc())
        .all()
    )


def ensure_company_access(db, user: User, company_id: Optional[str] = None) -> str:
    if is_admin(user):
        target_company_id = str(company_id or DEFAULT_COMPANY_ID)
        company = db.query(Company).filter(Company.id == target_company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail="Empresa nao encontrada.")
        return company.id

    memberships = active_company_memberships(db, user)
    if company_id:
        for membership in memberships:
            if membership.company_id == company_id:
                return membership.company_id
        raise HTTPException(status_code=403, detail="Usuario sem vinculo ativo com esta empresa.")

    if not memberships:
        raise HTTPException(status_code=403, detail="Usuario sem vinculo ativo com empresa.")
    return memberships[0].company_id


def get_company_membership(db, user: User, company_id: str) -> Optional[CompanyUser]:
    return (
        db.query(CompanyUser)
        .filter(
            CompanyUser.company_id == company_id,
            CompanyUser.user_id == user.id,
            CompanyUser.status == ACTIVE_STATUS,
        )
        .first()
    )


def is_company_admin(db, user: User, company_id: str) -> bool:
    if is_admin(user):
        return True
    membership = get_company_membership(db, user, company_id)
    return bool(membership and membership.role in TENANT_ADMIN_ROLES)


def ensure_company_admin(db, user: User, company_id: str):
    if not is_company_admin(db, user, company_id):
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador da empresa.")


def user_department_name(db, user: User, company_id: str) -> Optional[str]:
    membership = get_company_membership(db, user, company_id)
    if not membership or not membership.department_id:
        return user.department
    department = db.query(Department).filter(Department.id == membership.department_id).first()
    return department.name if department else user.department


def user_can_access_department(db, user: User, company_id: str, department_name: Optional[str]) -> bool:
    if is_company_admin(db, user, company_id):
        return True
    membership = get_company_membership(db, user, company_id)
    if not membership:
        return False
    if membership.role == "coordinator":
        return bool(department_name and user_department_name(db, user, company_id) == department_name)
    return False


def ensure_user_in_company(db, user_id: int, company_id: str) -> User:
    user = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not user:
        raise HTTPException(status_code=400, detail="Usuario nao encontrado.")
    if is_admin(user):
        return user
    membership = (
        db.query(CompanyUser)
        .filter(CompanyUser.company_id == company_id, CompanyUser.user_id == user.id, CompanyUser.status == ACTIVE_STATUS)
        .first()
    )
    if not membership:
        raise HTTPException(status_code=403, detail="Usuario nao pertence a esta empresa.")
    return user


def serialize_ticket(ticket: Ticket, db) -> dict:
    created_by = db.query(User).filter(User.id == ticket.created_by_id).first()
    assigned_to = db.query(User).filter(User.id == ticket.assigned_to_id).first() if ticket.assigned_to_id else None
    attachment = db.query(FileUpload).filter(FileUpload.id == ticket.attachment_file_id).first() if ticket.attachment_file_id else None
    first_response_minutes = None
    if ticket.created_at and ticket.first_response_at:
        first_response_minutes = max(0, int((ticket.first_response_at - ticket.created_at).total_seconds() // 60))

    return {
        "id": ticket.id,
        "company_id": ticket.company_id,
        "title": ticket.title,
        "description": ticket.description,
        "priority": ticket.priority,
        "status": ticket.status,
        "department": ticket.department,
        "channel": ticket.channel,
        "created_by_id": ticket.created_by_id,
        "created_by_name": created_by.full_name if created_by else None,
        "assigned_to_id": ticket.assigned_to_id,
        "assigned_to_name": assigned_to.full_name if assigned_to else None,
        "attachment_file_id": ticket.attachment_file_id,
        "attachment_filename": attachment.filename if attachment else None,
        "attachment_content_type": attachment.content_type if attachment else None,
        "attachment_file_size": attachment.file_size if attachment else None,
        "rating_score": ticket.rating_score,
        "rating_comment": ticket.rating_comment,
        "rated_at": ticket.rated_at.isoformat() if ticket.rated_at else None,
        "first_response_at": ticket.first_response_at.isoformat() if ticket.first_response_at else None,
        "first_response_minutes": first_response_minutes,
        "created_at": ticket.created_at.isoformat() if ticket.created_at else None,
        "updated_at": ticket.updated_at.isoformat() if ticket.updated_at else None,
        "closed_at": ticket.closed_at.isoformat() if ticket.closed_at else None,
    }


def serialize_ticket_message(message: TicketMessage, db) -> dict:
    sender = db.query(User).filter(User.id == message.sender_id).first()
    attachment = db.query(FileUpload).filter(FileUpload.id == message.file_id).first() if message.file_id else None
    return {
        "id": message.id,
        "company_id": message.company_id,
        "ticket_id": message.ticket_id,
        "sender_id": message.sender_id,
        "sender_name": sender.full_name if sender else "Usuario",
        "content": message.content,
        "file_id": message.file_id,
        "attachment_filename": attachment.filename if attachment else None,
        "attachment_content_type": attachment.content_type if attachment else None,
        "attachment_file_size": attachment.file_size if attachment else None,
        "created_at": message.created_at.isoformat() if message.created_at else None,
    }


def serialize_message_attachment(file_ref, db) -> dict:
    if not file_ref:
        return {}

    attachment = None
    parsed_file_id = None

    try:
        parsed_file_id = int(file_ref)
    except (TypeError, ValueError):
        parsed_file_id = None

    if parsed_file_id:
        attachment = db.query(FileUpload).filter(FileUpload.id == parsed_file_id).first()

    if not attachment:
        attachment = db.query(FileUpload).filter(FileUpload.file_path == str(file_ref)).first()

    if attachment:
        return {
            "file_id": attachment.id,
            "attachment_file_id": attachment.id,
            "attachment_filename": attachment.filename,
            "attachment_content_type": attachment.content_type,
            "attachment_file_size": attachment.file_size,
        }

    if parsed_file_id:
        return {
            "file_id": parsed_file_id,
            "attachment_file_id": parsed_file_id,
        }

    return {}


def serialize_message(message: Message, db) -> dict:
    sender = db.query(User).filter(User.id == message.sender_id).first()
    receiver = db.query(User).filter(User.id == message.receiver_id).first() if message.receiver_id else None
    return {
        "id": message.id,
        "company_id": message.company_id,
        "content": message.content,
        "sender_id": message.sender_id,
        "sender_name": sender.full_name if sender else "Usuario",
        "sender_department": sender.department if sender else None,
        "receiver_id": message.receiver_id,
        "receiver_name": receiver.full_name if receiver else None,
        "receiver_department": receiver.department if receiver else None,
        "message_type": message.message_type,
        "timestamp": message.timestamp.isoformat() if message.timestamp else None,
        "file_path": message.file_path,
        **serialize_message_attachment(message.file_path, db),
    }


def can_access_ticket(ticket: Ticket, user: User, db) -> bool:
    company_id = ticket.company_id or DEFAULT_COMPANY_ID
    try:
        ensure_company_access(db, user, company_id)
    except HTTPException:
        return False

    if is_company_admin(db, user, company_id):
        return True
    if user_can_access_department(db, user, company_id, ticket.department):
        return True
    return ticket.created_by_id == user.id or ticket.assigned_to_id == user.id


def ensure_ticket_access(ticket: Ticket, user: User, db):
    if not can_access_ticket(ticket, user, db):
        raise HTTPException(status_code=403, detail="Sem permissao para acessar este ticket.")


async def notify_ticket_user(user_id: Optional[int], title: str, description: str, ticket_id: int):
    if not user_id:
        return
    payload = {
        "type": "notification",
        "notification": {
            "id": f"ticket-{ticket_id}-{int(datetime.utcnow().timestamp() * 1000)}",
            "type": "Ticket",
            "title": title,
            "description": description,
            "time": "Agora",
            "unread": True,
            "ticket_id": ticket_id,
            "link": "/tickets",
        },
    }
    await manager.send_personal_message(json.dumps(payload), user_id)


def serialize_group(group: ChatGroup, db) -> dict:
    creator = db.query(User).filter(User.id == group.created_by_id).first()
    return {
        "id": group.id,
        "company_id": group.company_id,
        "name": group.name,
        "description": group.description,
        "department": group.department,
        "created_by_id": group.created_by_id,
        "created_by_name": creator.full_name if creator else None,
        "is_active": group.is_active,
        "created_at": group.created_at.isoformat() if group.created_at else None,
    }


def serialize_task(task: TaskItem, db) -> dict:
    created_by = db.query(User).filter(User.id == task.created_by_id).first()
    assigned_to = db.query(User).filter(User.id == task.assigned_to_id).first() if task.assigned_to_id else None
    return {
        "id": task.id,
        "company_id": task.company_id,
        "title": task.title,
        "description": task.description,
        "priority": task.priority,
        "category": task.category,
        "status": task.status,
        "due_date": task.due_date,
        "created_by_id": task.created_by_id,
        "created_by_name": created_by.full_name if created_by else None,
        "assigned_to_id": task.assigned_to_id,
        "assigned_to_name": assigned_to.full_name if assigned_to else None,
        "created_at": task.created_at.isoformat() if task.created_at else None,
        "updated_at": task.updated_at.isoformat() if task.updated_at else None,
    }


class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[int, WebSocket] = {}
        self.user_connections: Dict[int, int] = {}
        self.user_company_scope: Dict[int, str] = {}

    async def connect(self, websocket: WebSocket, user_id: int, company_id: str):
        await websocket.accept()
        connection_id = id(websocket)
        self.active_connections[connection_id] = websocket
        self.user_connections[user_id] = connection_id
        self.user_company_scope[user_id] = company_id
        await self.broadcast_user_status(user_id, True)
        await self.send_online_users(websocket, company_id)

    def disconnect(self, user_id: int):
        connection_id = self.user_connections.pop(user_id, None)
        self.user_company_scope.pop(user_id, None)
        if connection_id:
            self.active_connections.pop(connection_id, None)

    async def send_personal_message(self, message: str, user_id: int):
        connection_id = self.user_connections.get(user_id)
        websocket = self.active_connections.get(connection_id)
        if websocket:
            try:
                await websocket.send_text(message)
            except Exception:
                self.disconnect(user_id)

    async def broadcast(self, message: str, exclude_user: Optional[int] = None, company_id: Optional[str] = None):
        disconnected = []
        for user_id, connection_id in list(self.user_connections.items()):
            if exclude_user and user_id == exclude_user:
                continue
            if company_id and self.user_company_scope.get(user_id) != company_id:
                continue
            websocket = self.active_connections.get(connection_id)
            if websocket:
                try:
                    await websocket.send_text(message)
                except Exception:
                    disconnected.append(user_id)

        for user_id in disconnected:
            self.disconnect(user_id)

    async def broadcast_user_status(self, user_id: int, is_online: bool):
        with SessionLocal() as db:
            user = db.query(User).filter(User.id == user_id).first()
            if not user:
                return
            message = {
                "type": "user_status",
                "user_id": user_id,
                "username": user.username,
                "full_name": user.full_name,
                "is_online": is_online,
            }
            company_id = self.user_company_scope.get(user_id)
        await self.broadcast(json.dumps(message), exclude_user=user_id, company_id=company_id)

    async def send_online_users(self, websocket: WebSocket, company_id: str):
        online_users = []
        with SessionLocal() as db:
            for user_id in self.user_connections.keys():
                if self.user_company_scope.get(user_id) != company_id:
                    continue
                user = db.query(User).filter(User.id == user_id).first()
                if user:
                    online_users.append({"id": user.id, "username": user.username, "full_name": user.full_name})

        await websocket.send_text(json.dumps({"type": "online_users", "users": online_users}))


manager = ConnectionManager()


def create_default_users():
    default_users = [
        {
            "username": "admin",
            "email": "admin@voltcorp.com",
            "full_name": "Administrador Master",
            "password": "admin123",
            "access_level": "master",
            "department": "Administracao",
            "role_title": "Administrador do sistema",
            "phone_extension": "1000",
            "phone": "1000",
            "birthday": "01-01",
        },
        {
            "username": "coordenador",
            "email": "coord@voltcorp.com",
            "full_name": "Coordenador Sistema",
            "password": "coord123",
            "access_level": "coordenador",
            "department": "TI",
            "role_title": "Coordenador de TI",
            "phone_extension": "2000",
            "phone": "2000",
            "birthday": "02-02",
        },
        {
            "username": "usuario",
            "email": "user@voltcorp.com",
            "full_name": "Usuario Padrao",
            "password": "user123",
            "access_level": "usuario",
            "department": "TI",
            "role_title": "Analista",
            "phone_extension": "2001",
            "phone": "2001",
            "birthday": "03-03",
        },
    ]

    with SessionLocal() as db:
        for user_data in default_users:
            existing = db.query(User).filter(User.username == user_data["username"]).first()
            if existing:
                changed = False
                for field in ["department", "role_title", "phone_extension", "birthday"]:
                    if not getattr(existing, field, None):
                        setattr(existing, field, user_data[field])
                        changed = True
                if existing.access_level == "padrao":
                    existing.access_level = "usuario"
                    changed = True
                if existing.access_level == "master" and not existing.is_platform_admin:
                    existing.is_platform_admin = True
                    changed = True
                if not existing.status:
                    existing.status = ACTIVE_STATUS if existing.is_active else INACTIVE_STATUS
                    changed = True
                if not existing.phone:
                    existing.phone = user_data["phone"]
                    changed = True
                if changed:
                    existing.updated_at = datetime.utcnow()
                continue

            db.add(
                User(
                    username=user_data["username"],
                    uuid=str(uuid.uuid4()),
                    email=user_data["email"],
                    full_name=user_data["full_name"],
                    hashed_password=get_password_hash(user_data["password"]),
                    phone=user_data["phone"],
                    is_platform_admin=user_data["access_level"] == "master",
                    must_change_password=False,
                    status=ACTIVE_STATUS,
                    access_level=user_data["access_level"],
                    department=user_data["department"],
                    role_title=user_data["role_title"],
                    phone_extension=user_data["phone_extension"],
                    birthday=user_data["birthday"],
                )
            )
        db.commit()


def split_sql_statements(sql: str) -> list[str]:
    statements = []
    current = []
    in_single_quote = False
    in_double_quote = False

    for char in sql:
        if char == "'" and not in_double_quote:
            in_single_quote = not in_single_quote
        elif char == '"' and not in_single_quote:
            in_double_quote = not in_double_quote

        if char == ";" and not in_single_quote and not in_double_quote:
            statement = "".join(current).strip()
            if statement:
                statements.append(statement)
            current = []
            continue

        current.append(char)

    statement = "".join(current).strip()
    if statement:
        statements.append(statement)
    return statements


def normalize_text(value: str) -> str:
    replacements = {
        "á": "a",
        "à": "a",
        "ã": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    normalized = value.lower()
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    return normalized


def infer_priority(text: str, task_format: bool = False) -> str:
    normalized = normalize_text(text)
    if any(word in normalized for word in ["urgente", "critico", "critica", "alta", "prioridade maxima", "importante"]):
        return "high" if task_format else "Alta"
    if any(word in normalized for word in ["baixa", "sem pressa", "quando der", "normal baixa"]):
        return "low" if task_format else "Baixa"
    return "medium" if task_format else "Media"


def infer_due_date(text: str) -> Optional[str]:
    normalized = normalize_text(text)
    today = datetime.utcnow().date()
    if "amanha" in normalized:
        return (today + timedelta(days=1)).isoformat()
    if "hoje" in normalized:
        return today.isoformat()

    tokens = normalized.replace(",", " ").replace(".", " ").split()
    for token in tokens:
        parts = token.split("/")
        if len(parts) not in {2, 3}:
            continue
        try:
            day = int(parts[0])
            month = int(parts[1])
            year = int(parts[2]) if len(parts) == 3 else today.year
            return datetime(year, month, day).date().isoformat()
        except ValueError:
            continue
    return None


def infer_intent(text: str) -> str:
    normalized = normalize_text(text)
    task_words = ["tarefa", "task", "kanban", "atividade", "afazer", "a fazer", "pendencia"]
    ticket_words = ["ticket", "chamado", "solicitacao", "suporte", "problema", "incidente", "atendimento"]
    task_score = sum(1 for word in task_words if word in normalized)
    ticket_score = sum(1 for word in ticket_words if word in normalized)
    return "task" if task_score > ticket_score else "ticket"


def infer_department(text: str, users: list[User], fallback: Optional[str]) -> str:
    normalized = normalize_text(text)
    departments = {department for department in DEFAULT_DEPARTMENTS}
    departments.update(user.department for user in users if user.department)

    for department in sorted(departments, key=len, reverse=True):
        if normalize_text(department) in normalized:
            return department

    aliases = {
        "rh": "RH",
        "recursos humanos": "RH",
        "financeiro": "Financeiro",
        "financas": "Financeiro",
        "ti": "TI",
        "tecnologia": "TI",
        "suporte": "Suporte",
        "comercial": "Comercial",
        "operacao": "Operacao",
        "operacoes": "Operacao",
        "produto": "Produto",
    }
    for alias, department in aliases.items():
        if alias in normalized:
            return department

    return fallback or "Operacao"


def infer_assignee(text: str, users: list[User]) -> Optional[User]:
    normalized = normalize_text(text)
    for user in users:
        candidates = [user.username, user.email, user.full_name]
        if any(candidate and normalize_text(candidate) in normalized for candidate in candidates):
            return user
    return None


def extract_assistant_subject(text: str, intent: str) -> str:
    clean = " ".join(text.strip().split())
    prefixes = [
        "crie um ticket",
        "crie o ticket",
        "criar ticket",
        "abra um ticket",
        "abra o ticket",
        "abrir ticket",
        "registrar chamado",
        "registre chamado",
        "crie uma tarefa",
        "crie a tarefa",
        "criar tarefa",
        "adicione uma tarefa",
        "adicionar tarefa",
        "registrar",
        "registre",
    ]
    normalized = normalize_text(clean)
    subject = clean
    for prefix in prefixes:
        if normalized.startswith(prefix):
            subject = clean[len(prefix):].strip(" :-")
            break

    lowered = normalize_text(subject)
    if " sobre " in lowered:
        index = lowered.find(" sobre ")
        subject = subject[index + len(" sobre "):].strip(" :-")
    elif ":" in subject:
        subject = subject.split(":", 1)[1].strip()

    cut_markers = [" e atribua", " atribua", " ate amanha", " até amanhã", " para amanha", " para amanhã"]
    lowered = normalize_text(subject)
    for marker in cut_markers:
        index = lowered.find(normalize_text(marker))
        if index > 8:
            subject = subject[:index].strip(" .,;-")
            lowered = normalize_text(subject)

    filler = ["urgente", "alta prioridade", "prioridade alta", "por favor"]
    words = [word for word in subject.split() if normalize_text(word) not in filler]
    subject = " ".join(words).strip(" .,;-")
    if not subject:
        subject = "atendimento interno" if intent == "ticket" else "atividade operacional"
    return subject[:160]


def build_assistant_title(text: str, intent: str) -> str:
    subject = extract_assistant_subject(text, intent)
    words = subject.split()
    short_subject = " ".join(words[:9]).strip()
    if len(words) > 9:
        short_subject = f"{short_subject}..."
    prefix = "Ticket" if intent == "ticket" else "Tarefa"
    return f"{prefix}: {short_subject}".strip()[:180]


def build_assistant_scope(subject: str, intent: str, department: str, assignee: Optional[User], due_date: Optional[str]) -> str:
    owner = assignee.full_name if assignee else "responsavel definido pelo setor"
    due_text = f" Prazo identificado: {due_date}." if due_date else ""
    if intent == "ticket":
        return (
            f"Escopo: analisar {subject}, identificar causa, registrar evidencias e conduzir o atendimento ate a "
            f"resolucao ou proximo encaminhamento. Setor: {department}. Responsavel: {owner}.{due_text}"
        )
    return (
        f"Escopo: executar {subject}, organizar as etapas necessarias, registrar andamento no Kanban e concluir com "
        f"validacao do responsavel. Setor: {department}. Responsavel: {owner}.{due_text}"
    )


def plan_assistant_action(text: str, current_user: User, db, company_id: str) -> dict:
    users = (
        db.query(User)
        .join(CompanyUser, CompanyUser.user_id == User.id)
        .filter(User.is_active == True, CompanyUser.company_id == company_id, CompanyUser.status == ACTIVE_STATUS)
        .all()
    )
    intent = infer_intent(text)
    assignee = infer_assignee(text, users)
    department = infer_department(text, users, current_user.department)
    due_date = infer_due_date(text)
    subject = extract_assistant_subject(text, intent)
    title = build_assistant_title(subject, intent)

    return {
        "intent": intent,
        "title": title,
        "description": build_assistant_scope(subject, intent, department, assignee, due_date),
        "subject": subject,
        "department": department,
        "assigned_to_id": assignee.id if assignee else None,
        "assigned_to_name": assignee.full_name if assignee else None,
        "ticket_priority": infer_priority(text, task_format=False),
        "task_priority": infer_priority(text, task_format=True),
        "due_date": due_date,
    }


def run_startup_migrations():
    if DATABASE_URL.startswith("sqlite"):
        return
    if os.getenv("AUTO_MIGRATE_DB", "true").lower() != "true":
        return

    migrations_dir = Path(__file__).resolve().parent / "db" / "migrations"
    if not migrations_dir.exists():
        return

    with engine.begin() as conn:
        conn.exec_driver_sql(
            """
            CREATE TABLE IF NOT EXISTS schema_migrations (
              version TEXT PRIMARY KEY,
              applied_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )

        for migration in sorted(migrations_dir.glob("*.sql")):
            sql = migration.read_text(encoding="utf-8")
            for statement in split_sql_statements(sql):
                conn.exec_driver_sql(statement)
            conn.exec_driver_sql(
                "INSERT INTO schema_migrations (version) VALUES (%s) ON CONFLICT (version) DO NOTHING",
                (migration.stem,),
            )


@app.on_event("startup")
async def startup_event():
    run_startup_migrations()
    if DATABASE_URL.startswith("sqlite") or os.getenv("AUTO_SEED_USERS", "false").lower() == "true":
        create_default_users()
    ensure_default_tenant_records()


@app.get("/")
async def root():
    return {"message": "Volt Corp API", "version": "1.0.0", "status": "online"}


@app.get("/health")
async def health_check():
    return {"status": "healthy", "database": "postgres" if "postgresql" in DATABASE_URL else "sqlite"}


@app.get("/version")
async def version_check():
    return {
        "service": "api",
        "app": "Volt Corp",
        "version": APP_VERSION,
        "commit": os.getenv("RENDER_GIT_COMMIT") or APP_VERSION,
        "build_time": APP_BUILD_TIME,
    }


@app.get("/downloads/desktop/latest/meta")
async def latest_desktop_release_meta():
    with SessionLocal() as db:
        release = (
            db.query(DesktopRelease)
            .filter(DesktopRelease.platform == "windows", DesktopRelease.is_active == True)
            .order_by(DesktopRelease.created_at.desc())
            .first()
        )
        if not release:
            raise HTTPException(status_code=404, detail="Instalador desktop ainda nao publicado")

        return {
            "version": release.version,
            "platform": release.platform,
            "filename": release.filename,
            "content_type": release.content_type,
            "file_size": release.file_size,
            "sha256": release.sha256,
            "storage_mode": release.storage_mode,
            "created_at": release.created_at.isoformat() if release.created_at else None,
            "download_url": "/downloads/desktop/latest",
        }


def iter_desktop_release_chunks(release_id: int):
    with SessionLocal() as db:
        chunk_index = 0
        while True:
            chunk = (
                db.query(DesktopReleaseChunk)
                .filter(
                    DesktopReleaseChunk.release_id == release_id,
                    DesktopReleaseChunk.chunk_index == chunk_index,
                )
                .first()
            )
            if not chunk:
                break
            yield bytes(chunk.data)
            chunk_index += 1


@app.get("/downloads/desktop/latest")
async def download_latest_desktop_release():
    with SessionLocal() as db:
        release = (
            db.query(DesktopRelease)
            .filter(DesktopRelease.platform == "windows", DesktopRelease.is_active == True)
            .order_by(DesktopRelease.created_at.desc())
            .first()
        )
        if not release:
            raise HTTPException(status_code=404, detail="Instalador desktop ainda nao publicado")

        release_id = release.id
        storage_mode = release.storage_mode or "inline"
        content_type = release.content_type
        payload = bytes(release.binary_data or b"") if storage_mode == "inline" else None
        chunk_count = 0
        if storage_mode == "chunks":
            try:
                chunk_count = (
                    db.query(DesktopReleaseChunk)
                    .filter(DesktopReleaseChunk.release_id == release_id)
                    .count()
                )
            except Exception as exc:
                print(f"Erro ao consultar chunks do instalador: {exc}")
                raise HTTPException(status_code=503, detail="Armazenamento do instalador ainda nao esta pronto.")
        headers = {
            "Content-Disposition": f'attachment; filename="{release.filename}"',
            "Content-Length": str(release.file_size),
            "Cache-Control": "no-store",
            "X-VoltCorp-Version": release.version,
            "X-VoltCorp-Sha256": release.sha256,
        }

    if storage_mode == "chunks":
        if chunk_count == 0:
            raise HTTPException(status_code=404, detail="Instalador desktop sem dados publicados")
        return StreamingResponse(
            iter_desktop_release_chunks(release_id),
            media_type=content_type,
            headers=headers,
        )

    if not payload:
        raise HTTPException(status_code=404, detail="Instalador desktop sem arquivo publicado")

    return StreamingResponse(iter([payload]), media_type=content_type, headers=headers)


@app.post("/auth/login")
async def login(credentials: dict):
    username = credentials.get("username")
    password = credentials.get("password")

    if not username or not password:
        raise HTTPException(status_code=400, detail="Username e password sao obrigatorios")

    with SessionLocal() as db:
        user = db.query(User).filter(or_(User.username == username, User.email == username)).first()
        if not user or not verify_password(password, user.hashed_password):
            raise HTTPException(status_code=401, detail="Credenciais invalidas")
        if not user.is_active or normalize_status(getattr(user, "status", ACTIVE_STATUS)) != ACTIVE_STATUS:
            raise HTTPException(status_code=401, detail="Usuario inativo")

        access_token = create_access_token(
            data={"sub": str(user.id)},
            expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
        )
        return {
            "access_token": access_token,
            "token_type": "bearer",
            "user": serialize_user(user, db),
        }


@app.post("/auth/logout")
async def logout(current_user: User = Depends(get_current_user)):
    return {"message": "Logout realizado com sucesso"}


@app.get("/auth/me")
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        user = db.query(User).filter(User.id == current_user.id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuario nao encontrado.")
        return serialize_user(user, db)


@app.post("/auth/change-password")
async def change_password(payload: dict, current_user: User = Depends(get_current_user)):
    new_password = str(payload.get("new_password") or "").strip()
    current_password = str(payload.get("current_password") or "").strip()
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="A nova senha deve ter pelo menos 6 caracteres.")

    with SessionLocal() as db:
        user = db.query(User).filter(User.id == current_user.id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuario nao encontrado.")
        if current_password and not verify_password(current_password, user.hashed_password):
            raise HTTPException(status_code=401, detail="Senha atual invalida.")

        user.hashed_password = get_password_hash(new_password)
        user.must_change_password = False
        user.updated_at = datetime.utcnow()
        log_audit(
            db,
            user.id,
            "senha_alterada",
            "user",
            user.id,
            ensure_company_access(db, user) if not is_admin(user) else None,
            {"self_service": True},
        )
        db.commit()
        db.refresh(user)
        return {"message": "Senha alterada com sucesso.", "user": serialize_user(user, db)}


@app.get("/users/")
async def list_users(company_id: Optional[str] = Query(default=None), current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        query = db.query(User)
        if is_admin(current_user) and not company_id:
            users = query.order_by(User.full_name.asc()).all()
            return [serialize_user(item, db) for item in users]

        scoped_company_id = ensure_company_access(db, current_user, company_id)
        query = query.join(CompanyUser, CompanyUser.user_id == User.id).filter(
            CompanyUser.company_id == scoped_company_id,
            CompanyUser.status == ACTIVE_STATUS,
        )
        if not is_company_admin(db, current_user, scoped_company_id):
            membership = get_company_membership(db, current_user, scoped_company_id)
            if membership and membership.role == "coordinator" and membership.department_id:
                query = query.filter(CompanyUser.department_id == membership.department_id)
            else:
                query = query.filter(User.id == current_user.id)
        users = query.order_by(User.full_name.asc()).all()
        return [serialize_user(item, db) for item in users]


@app.get("/birthdays/")
async def list_birthdays(current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user)
        users = (
            db.query(User)
            .join(CompanyUser, CompanyUser.user_id == User.id)
            .filter(User.is_active == True, User.birthday.isnot(None), User.birthday != "")
            .filter(CompanyUser.company_id == company_id, CompanyUser.status == ACTIVE_STATUS)
            .order_by(User.full_name.asc())
            .all()
        )
        return [serialize_user(item, db) for item in users]


@app.post("/users/")
async def create_user(payload: dict, current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user, payload.get("company_id"))
        membership = get_company_membership(db, current_user, company_id)
        can_create = is_admin(current_user) or (membership and membership.role in {"company_admin", "master_admin", "coordinator"})
    if not can_create:
        raise HTTPException(status_code=403, detail="Sem permissao para criar usuarios.")

    required_fields = ["email", "password"]
    missing = [field for field in required_fields if not str(payload.get(field) or "").strip()]
    if missing:
        raise HTTPException(status_code=400, detail=f"Campos obrigatorios: {', '.join(missing)}")

    email = str(payload["email"]).strip().lower()
    if not EMAIL_PATTERN.match(email):
        raise HTTPException(status_code=400, detail="Email invalido.")

    role = normalize_platform_role(payload.get("role") or payload.get("access_level") or "user")
    if role == "master_admin" and not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Somente Admin Master cria administradores globais.")

    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user, payload.get("company_id"))
        ensure_company_admin(db, current_user, company_id) if role in {"company_admin", "coordinator"} else None
        department_id = payload.get("department_id")
        if department_id:
            department = db.query(Department).filter(Department.id == str(department_id), Department.company_id == company_id).first()
            if not department:
                raise HTTPException(status_code=400, detail="Setor nao encontrado.")
        else:
            department = get_or_create_department(db, company_id, str(payload.get("department") or current_user.department or "Operacao"))

        if not is_company_admin(db, current_user, company_id):
            current_membership = get_company_membership(db, current_user, company_id)
            if not current_membership or current_membership.role != "coordinator" or current_membership.department_id != department.id:
                raise HTTPException(status_code=403, detail="Coordenador so pode criar usuarios do proprio setor.")
            role = "user"

        username = str(payload.get("username") or email.split("@")[0]).strip().lower()
        duplicate = db.query(User).filter(or_(User.username == username, User.email == email)).first()
        if duplicate:
            user = duplicate
        else:
            user = User(
                uuid=str(uuid.uuid4()),
                username=username,
                email=email,
                full_name=str(payload.get("full_name") or payload.get("name") or email).strip(),
                hashed_password=get_password_hash(str(payload["password"])),
                phone=str(payload.get("phone") or payload.get("phone_extension") or "").strip() or None,
                must_change_password=True,
                status=normalize_status(payload.get("status")),
                is_active=normalize_status(payload.get("status")) == ACTIVE_STATUS,
                is_platform_admin=role == "master_admin",
                access_level="master" if role == "master_admin" else legacy_access_for_role(role),
                department=department.name,
                phone_extension=str(payload.get("phone_extension") or payload.get("phone") or "").strip() or None,
                birthday=str(payload.get("birthday") or "").strip() or None,
                role_title=str(payload.get("role_title") or "").strip() or None,
            )
            db.add(user)
            db.flush()
            log_audit(db, current_user.id, "usuario_criado", "user", user.id, company_id, {"email": user.email, "role": role})

        existing_link = (
            db.query(CompanyUser)
            .filter(CompanyUser.company_id == company_id, CompanyUser.user_id == user.id)
            .first()
        )
        if existing_link:
            existing_link.department_id = department.id
            existing_link.role = role if role != "master_admin" else "company_admin"
            existing_link.status = normalize_status(payload.get("status"))
        else:
            db.add(
                CompanyUser(
                    id=str(uuid.uuid4()),
                    company_id=company_id,
                    user_id=user.id,
                    department_id=department.id,
                    role=role if role != "master_admin" else "company_admin",
                    status=normalize_status(payload.get("status")),
                )
            )
            log_audit(db, current_user.id, "vinculo_criado", "company_user", user.id, company_id, {"role": role})
        db.commit()
        db.refresh(user)
        return serialize_user(user, db)


@app.put("/users/{user_id}")
async def update_user(user_id: int, payload: dict, current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        target = db.query(User).filter(User.id == user_id).first()
        if not target:
            raise HTTPException(status_code=404, detail="Usuario nao encontrado.")

        company_id = ensure_company_access(db, current_user, payload.get("company_id"))
        target_membership = (
            db.query(CompanyUser)
            .filter(CompanyUser.company_id == company_id, CompanyUser.user_id == target.id)
            .first()
        )
        if not is_admin(current_user):
            if not target_membership:
                raise HTTPException(status_code=403, detail="Usuario nao pertence a esta empresa.")
            if not is_company_admin(db, current_user, company_id):
                current_membership = get_company_membership(db, current_user, company_id)
                if not current_membership or current_membership.role != "coordinator" or target_membership.department_id != current_membership.department_id:
                    raise HTTPException(status_code=403, detail="Sem permissao para editar este usuario.")

        editable_fields = ["full_name", "email", "phone", "phone_extension", "birthday", "role_title"]
        for field in editable_fields:
            if field in payload:
                value = str(payload.get(field) or "").strip() or None
                setattr(target, field, value)

        if "department_id" in payload and target_membership:
            department = db.query(Department).filter(Department.id == str(payload.get("department_id")), Department.company_id == company_id).first()
            if not department:
                raise HTTPException(status_code=400, detail="Setor nao encontrado.")
            target_membership.department_id = department.id
            target.department = department.name

        if "role" in payload and target_membership:
            role = normalize_platform_role(payload.get("role"))
            if role == "master_admin" and not is_admin(current_user):
                raise HTTPException(status_code=403, detail="Somente Admin Master altera administrador global.")
            target_membership.role = "company_admin" if role == "master_admin" else role
            target.access_level = "master" if role == "master_admin" else legacy_access_for_role(role)
            target.is_platform_admin = role == "master_admin"
            log_audit(db, current_user.id, "nivel_alterado", "company_user", target_membership.id, company_id, {"role": target_membership.role})

        if "is_active" in payload and is_admin(current_user):
            target.is_active = bool(payload.get("is_active"))
            target.status = ACTIVE_STATUS if target.is_active else INACTIVE_STATUS
        if "status" in payload:
            target.status = normalize_status(payload.get("status"))
            target.is_active = target.status == ACTIVE_STATUS
            if target_membership:
                target_membership.status = target.status

        target.updated_at = datetime.utcnow()
        log_audit(db, current_user.id, "usuario_editado", "user", target.id, company_id, {"email": target.email})
        db.commit()
        db.refresh(target)
        return serialize_user(target, db)


@app.get("/departments/")
async def list_departments(current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user)
        query = db.query(Department).filter(Department.company_id == company_id, Department.status == ACTIVE_STATUS)
        if not is_company_admin(db, current_user, company_id):
            membership = get_company_membership(db, current_user, company_id)
            if membership and membership.department_id:
                query = query.filter(Department.id == membership.department_id)
        return [department.name for department in query.order_by(Department.name.asc()).all()]


@app.get("/tickets/")
async def list_tickets(
    status: Optional[str] = Query(default=None),
    department: Optional[str] = Query(default=None),
    company_id: Optional[str] = Query(default=None),
    limit: int = Query(default=100, ge=1, le=200),
    current_user: User = Depends(get_current_user),
):
    with SessionLocal() as db:
        scoped_company_id = ensure_company_access(db, current_user, company_id)
        query = db.query(Ticket).filter(Ticket.company_id == scoped_company_id)
        if is_company_admin(db, current_user, scoped_company_id):
            if department:
                query = query.filter(Ticket.department == department)
        elif user_can_access_department(db, current_user, scoped_company_id, user_department_name(db, current_user, scoped_company_id)):
            query = query.filter(Ticket.department == user_department_name(db, current_user, scoped_company_id))
        else:
            query = query.filter(or_(Ticket.created_by_id == current_user.id, Ticket.assigned_to_id == current_user.id))

        if status:
            query = query.filter(Ticket.status == status)

        tickets = query.order_by(Ticket.created_at.desc()).limit(limit).all()
        return [serialize_ticket(ticket, db) for ticket in tickets]


@app.post("/tickets/")
async def create_ticket(payload: dict, current_user: User = Depends(get_current_user)):
    title = str(payload.get("title") or "").strip()
    description = str(payload.get("description") or "").strip()
    if not title or not description:
        raise HTTPException(status_code=400, detail="Titulo e descricao sao obrigatorios.")

    assigned_to_id = payload.get("assigned_to_id")
    attachment_file_id = payload.get("attachment_file_id")
    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user, payload.get("company_id"))
        department = str(payload.get("department") or user_department_name(db, current_user, company_id) or "Operacao").strip()
        get_or_create_department(db, company_id, department)
        if not is_company_admin(db, current_user, company_id) and not user_can_access_department(db, current_user, company_id, department):
            raise HTTPException(status_code=403, detail="Coordenador so pode criar tickets do proprio setor.")
        if assigned_to_id:
            assigned_user = ensure_user_in_company(db, int(assigned_to_id), company_id)
            if not is_company_admin(db, current_user, company_id) and assigned_user.department != department:
                raise HTTPException(status_code=403, detail="Responsavel precisa estar no mesmo setor.")
        if attachment_file_id:
            attachment = db.query(FileUpload).filter(FileUpload.id == int(attachment_file_id), FileUpload.company_id == company_id).first()
            if not attachment:
                raise HTTPException(status_code=400, detail="Anexo nao encontrado.")

        ticket = Ticket(
            company_id=company_id,
            title=title,
            description=description,
            priority=str(payload.get("priority") or "Media"),
            status=str(payload.get("status") or "Aberto"),
            department=department,
            channel=str(payload.get("channel") or "Web"),
            created_by_id=current_user.id,
            assigned_to_id=int(assigned_to_id) if assigned_to_id else None,
            attachment_file_id=int(attachment_file_id) if attachment_file_id else None,
        )
        db.add(ticket)
        db.commit()
        db.refresh(ticket)
        ticket_response = serialize_ticket(ticket, db)

        if ticket.assigned_to_id and ticket.assigned_to_id != current_user.id:
            await notify_ticket_user(
                ticket.assigned_to_id,
                "Novo ticket atribuido",
                f"{current_user.full_name} abriu: {ticket.title}",
                ticket.id,
            )

        return ticket_response


@app.get("/tickets/{ticket_id}")
async def get_ticket(ticket_id: int, current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket nao encontrado.")
        ensure_ticket_access(ticket, current_user, db)
        return serialize_ticket(ticket, db)


@app.get("/tickets/{ticket_id}/messages")
async def list_ticket_messages(ticket_id: int, current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket nao encontrado.")
        ensure_ticket_access(ticket, current_user, db)

        messages = (
            db.query(TicketMessage)
            .filter(TicketMessage.ticket_id == ticket_id)
            .order_by(TicketMessage.created_at.asc())
            .all()
        )
        return [serialize_ticket_message(message, db) for message in messages]


@app.post("/tickets/{ticket_id}/messages")
async def create_ticket_message(ticket_id: int, payload: dict, current_user: User = Depends(get_current_user)):
    content = str(payload.get("content") or "").strip()
    file_id = payload.get("file_id")
    if not content and not file_id:
        raise HTTPException(status_code=400, detail="Informe uma mensagem ou anexo.")

    with SessionLocal() as db:
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket nao encontrado.")
        ensure_ticket_access(ticket, current_user, db)
        if ticket.status == "Resolvido":
            raise HTTPException(status_code=400, detail="Ticket resolvido nao recebe novas mensagens.")

        if file_id:
            attachment = db.query(FileUpload).filter(FileUpload.id == int(file_id), FileUpload.company_id == ticket.company_id).first()
            if not attachment:
                raise HTTPException(status_code=400, detail="Anexo nao encontrado.")

        message = TicketMessage(
            company_id=ticket.company_id,
            ticket_id=ticket.id,
            sender_id=current_user.id,
            content=content,
            file_id=int(file_id) if file_id else None,
        )
        db.add(message)

        if not ticket.first_response_at and current_user.id != ticket.created_by_id:
            ticket.first_response_at = datetime.utcnow()
        if ticket.status == "Aberto" and current_user.id != ticket.created_by_id:
            ticket.status = "Em andamento"
        ticket.updated_at = datetime.utcnow()

        db.commit()
        db.refresh(message)
        db.refresh(ticket)
        message_response = serialize_ticket_message(message, db)

        notify_user_id = ticket.created_by_id if current_user.id == ticket.assigned_to_id else ticket.assigned_to_id
        if notify_user_id and notify_user_id != current_user.id:
            await notify_ticket_user(
                notify_user_id,
                "Nova conversa no ticket",
                f"{current_user.full_name} respondeu em: {ticket.title}",
                ticket.id,
            )

        return message_response


@app.patch("/tickets/{ticket_id}/transfer")
async def transfer_ticket(ticket_id: int, payload: dict, current_user: User = Depends(get_current_user)):
    assigned_to_id = payload.get("assigned_to_id")
    if not assigned_to_id:
        raise HTTPException(status_code=400, detail="Selecione um novo responsavel.")

    note = str(payload.get("message") or "").strip()
    with SessionLocal() as db:
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket nao encontrado.")
        ensure_ticket_access(ticket, current_user, db)

        assigned_user = ensure_user_in_company(db, int(assigned_to_id), ticket.company_id)
        if not is_company_admin(db, current_user, ticket.company_id) and assigned_user.department != ticket.department:
            raise HTTPException(status_code=403, detail="Responsavel precisa estar no mesmo setor.")

        ticket.assigned_to_id = assigned_user.id
        ticket.status = "Em andamento" if ticket.status != "Resolvido" else ticket.status
        ticket.updated_at = datetime.utcnow()
        db.add(TicketMessage(
            company_id=ticket.company_id,
            ticket_id=ticket.id,
            sender_id=current_user.id,
            content=note or f"Ticket repassado para {assigned_user.full_name}.",
        ))
        db.commit()
        db.refresh(ticket)
        ticket_response = serialize_ticket(ticket, db)

        if assigned_user.id != current_user.id:
            await notify_ticket_user(
                assigned_user.id,
                "Ticket repassado para voce",
                f"{current_user.full_name} repassou: {ticket.title}",
                ticket.id,
            )

        return ticket_response


@app.post("/tickets/{ticket_id}/close")
async def close_ticket(ticket_id: int, payload: dict, current_user: User = Depends(get_current_user)):
    note = str(payload.get("message") or "").strip()
    with SessionLocal() as db:
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket nao encontrado.")
        ensure_ticket_access(ticket, current_user, db)

        ticket.status = "Resolvido"
        ticket.closed_at = ticket.closed_at or datetime.utcnow()
        ticket.updated_at = datetime.utcnow()
        db.add(TicketMessage(
            company_id=ticket.company_id,
            ticket_id=ticket.id,
            sender_id=current_user.id,
            content=note or "Ticket fechado.",
        ))
        db.commit()
        db.refresh(ticket)
        ticket_response = serialize_ticket(ticket, db)

        if ticket.created_by_id != current_user.id:
            await notify_ticket_user(
                ticket.created_by_id,
                "Ticket fechado",
                f"{current_user.full_name} fechou: {ticket.title}. Avalie o atendimento.",
                ticket.id,
            )

        return ticket_response


@app.post("/tickets/{ticket_id}/rating")
async def rate_ticket(ticket_id: int, payload: dict, current_user: User = Depends(get_current_user)):
    try:
        rating_score = int(payload.get("rating_score"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Informe uma avaliacao de 1 a 5.")
    if rating_score < 1 or rating_score > 5:
        raise HTTPException(status_code=400, detail="A avaliacao deve ficar entre 1 e 5.")

    with SessionLocal() as db:
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket nao encontrado.")
        ensure_ticket_access(ticket, current_user, db)
        if ticket.created_by_id != current_user.id:
            raise HTTPException(status_code=403, detail="Somente o dono do ticket pode avaliar.")
        if ticket.status != "Resolvido":
            raise HTTPException(status_code=400, detail="Avalie apenas depois do fechamento.")

        ticket.rating_score = rating_score
        ticket.rating_comment = str(payload.get("rating_comment") or "").strip() or None
        ticket.rated_at = datetime.utcnow()
        ticket.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(ticket)
        return serialize_ticket(ticket, db)


@app.patch("/tickets/{ticket_id}")
async def update_ticket(ticket_id: int, payload: dict, current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket nao encontrado.")

        ensure_company_access(db, current_user, ticket.company_id)
        can_edit = is_company_admin(db, current_user, ticket.company_id)
        if not can_edit:
            can_edit = user_can_access_department(db, current_user, ticket.company_id, ticket.department)
        if not can_edit:
            can_edit = ticket.created_by_id == current_user.id or ticket.assigned_to_id == current_user.id
        if not can_edit:
            raise HTTPException(status_code=403, detail="Sem permissao para editar este ticket.")

        previous_assigned_to_id = ticket.assigned_to_id

        for field in ["title", "description", "priority", "status", "department", "channel"]:
            if field in payload:
                value = str(payload.get(field) or "").strip() or None
                if field == "department" and value:
                    get_or_create_department(db, ticket.company_id, value)
                    if not is_company_admin(db, current_user, ticket.company_id) and not user_can_access_department(db, current_user, ticket.company_id, value):
                        raise HTTPException(status_code=403, detail="Sem permissao para mover ticket para outro setor.")
                setattr(ticket, field, value)

        if "assigned_to_id" in payload:
            assigned_to_id = payload.get("assigned_to_id")
            if assigned_to_id:
                assigned_user = ensure_user_in_company(db, int(assigned_to_id), ticket.company_id)
                if not is_company_admin(db, current_user, ticket.company_id) and assigned_user.department != ticket.department:
                    raise HTTPException(status_code=403, detail="Responsavel precisa estar no mesmo setor.")
                ticket.assigned_to_id = assigned_user.id
            else:
                ticket.assigned_to_id = None

        if "attachment_file_id" in payload:
            attachment_file_id = payload.get("attachment_file_id")
            if attachment_file_id:
                attachment = db.query(FileUpload).filter(FileUpload.id == int(attachment_file_id), FileUpload.company_id == ticket.company_id).first()
                if not attachment:
                    raise HTTPException(status_code=400, detail="Anexo nao encontrado.")
                ticket.attachment_file_id = attachment.id
            else:
                ticket.attachment_file_id = None

        if ticket.status == "Resolvido":
            ticket.closed_at = ticket.closed_at or datetime.utcnow()
        else:
            ticket.closed_at = None
        ticket.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(ticket)
        ticket_response = serialize_ticket(ticket, db)

        if ticket.assigned_to_id and ticket.assigned_to_id != previous_assigned_to_id and ticket.assigned_to_id != current_user.id:
            await notify_ticket_user(
                ticket.assigned_to_id,
                "Ticket atribuido a voce",
                f"{current_user.full_name} atribuiu: {ticket.title}",
                ticket.id,
            )

        return ticket_response


@app.get("/tasks/")
async def list_tasks(
    status: Optional[str] = Query(default=None),
    company_id: Optional[str] = Query(default=None),
    current_user: User = Depends(get_current_user),
):
    with SessionLocal() as db:
        scoped_company_id = ensure_company_access(db, current_user, company_id)
        query = db.query(TaskItem).filter(TaskItem.company_id == scoped_company_id)
        if status:
            query = query.filter(TaskItem.status == status)
        if not is_company_admin(db, current_user, scoped_company_id):
            department_name = user_department_name(db, current_user, scoped_company_id)
            if user_can_access_department(db, current_user, scoped_company_id, department_name):
                query = query.filter(TaskItem.category == department_name)
            else:
                query = query.filter(or_(TaskItem.created_by_id == current_user.id, TaskItem.assigned_to_id == current_user.id))
        tasks = query.order_by(TaskItem.created_at.desc()).all()
        return [serialize_task(task, db) for task in tasks]


@app.post("/tasks/")
async def create_task(payload: dict, current_user: User = Depends(get_current_user)):
    title = str(payload.get("title") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Titulo da tarefa e obrigatorio.")

    assigned_to_id = payload.get("assigned_to_id")

    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user, payload.get("company_id"))
        category = str(payload.get("category") or user_department_name(db, current_user, company_id) or "Operacao").strip()
        get_or_create_department(db, company_id, category)
        if not is_company_admin(db, current_user, company_id) and not user_can_access_department(db, current_user, company_id, category):
            raise HTTPException(status_code=403, detail="Coordenador so pode criar tarefas do proprio setor.")
        assigned_user = None
        if assigned_to_id:
            assigned_user = ensure_user_in_company(db, int(assigned_to_id), company_id)
            if not is_company_admin(db, current_user, company_id) and assigned_user.department != category:
                raise HTTPException(status_code=403, detail="Responsavel precisa estar no mesmo setor.")

        task = TaskItem(
            company_id=company_id,
            title=title,
            description=str(payload.get("description") or "").strip() or None,
            priority=str(payload.get("priority") or "medium").strip(),
            category=category,
            status=str(payload.get("status") or "backlog").strip(),
            due_date=str(payload.get("due_date") or "").strip() or None,
            created_by_id=current_user.id,
            assigned_to_id=assigned_user.id if assigned_user else None,
        )
        db.add(task)
        db.commit()
        db.refresh(task)
        return serialize_task(task, db)


@app.patch("/tasks/{task_id}")
async def update_task(task_id: int, payload: dict, current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        task = db.query(TaskItem).filter(TaskItem.id == task_id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Tarefa nao encontrada.")

        ensure_company_access(db, current_user, task.company_id)
        can_edit = is_company_admin(db, current_user, task.company_id) or task.created_by_id == current_user.id or task.assigned_to_id == current_user.id
        if not can_edit:
            can_edit = user_can_access_department(db, current_user, task.company_id, task.category)
        if not can_edit:
            raise HTTPException(status_code=403, detail="Sem permissao para editar esta tarefa.")

        for field in ["title", "description", "priority", "category", "status", "due_date"]:
            if field in payload:
                value = str(payload.get(field) or "").strip() or None
                if field in ["title", "priority", "category", "status"]:
                    value = value or getattr(task, field)
                if field == "category":
                    get_or_create_department(db, task.company_id, value)
                    if not is_company_admin(db, current_user, task.company_id) and not user_can_access_department(db, current_user, task.company_id, value):
                        raise HTTPException(status_code=403, detail="Sem permissao para mover tarefa para outro setor.")
                setattr(task, field, value)

        if "assigned_to_id" in payload:
            assigned_to_id = payload.get("assigned_to_id")
            if assigned_to_id:
                assigned_user = ensure_user_in_company(db, int(assigned_to_id), task.company_id)
                task.assigned_to_id = assigned_user.id
            else:
                task.assigned_to_id = None

        task.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(task)
        return serialize_task(task, db)


@app.delete("/tasks/{task_id}")
async def delete_task(task_id: int, current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        task = db.query(TaskItem).filter(TaskItem.id == task_id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Tarefa nao encontrada.")
        ensure_company_access(db, current_user, task.company_id)
        if not is_company_admin(db, current_user, task.company_id) and task.created_by_id != current_user.id:
            raise HTTPException(status_code=403, detail="Sem permissao para remover esta tarefa.")
        db.delete(task)
        db.commit()
        return {"message": "Tarefa removida."}


@app.post("/assistant/requests")
async def assistant_request(payload: dict, current_user: User = Depends(get_current_user)):
    message = str(payload.get("message") or "").strip()
    execute = bool(payload.get("execute", True))
    if not message:
        raise HTTPException(status_code=400, detail="Informe uma solicitacao para o assistente.")

    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user, payload.get("company_id"))
        plan = plan_assistant_action(message, current_user, db, company_id)
        if not is_company_admin(db, current_user, company_id) and not user_can_access_department(db, current_user, company_id, plan["department"]):
            raise HTTPException(status_code=403, detail="Coordenador so pode criar itens do proprio setor.")

        response = {
            "intent": plan["intent"],
            "plan": plan,
            "executed": False,
            "ticket": None,
            "task": None,
            "reply": "",
        }

        if not execute:
            response["reply"] = "Plano montado. Confirme para criar no Volt Corp."
            return response

        if plan["intent"] == "ticket":
            ticket = Ticket(
                company_id=company_id,
                title=plan["title"],
                description=plan["description"],
                priority=plan["ticket_priority"],
                status="Aberto",
                department=plan["department"],
                channel="Assistente",
                created_by_id=current_user.id,
                assigned_to_id=plan["assigned_to_id"],
            )
            db.add(ticket)
            db.commit()
            db.refresh(ticket)
            response["executed"] = True
            response["ticket"] = serialize_ticket(ticket, db)
            response["reply"] = f"Criei o ticket #{ticket.id} para {plan['department']} com prioridade {plan['ticket_priority']}."
            if ticket.assigned_to_id and ticket.assigned_to_id != current_user.id:
                await notify_ticket_user(
                    ticket.assigned_to_id,
                    "Novo ticket atribuido",
                    f"{current_user.full_name} abriu via assistente: {ticket.title}",
                    ticket.id,
                )
            return response

        task = TaskItem(
            company_id=company_id,
            title=plan["title"],
            description=plan["description"],
            priority=plan["task_priority"],
            category=plan["department"],
            status="backlog",
            due_date=plan["due_date"],
            created_by_id=current_user.id,
            assigned_to_id=plan["assigned_to_id"],
        )
        db.add(task)
        db.commit()
        db.refresh(task)
        response["executed"] = True
        response["task"] = serialize_task(task, db)
        response["reply"] = f"Criei a tarefa #{task.id} no Kanban em {plan['department']}."
        return response


async def read_import_rows(file: UploadFile) -> list[dict]:
    content = await file.read()
    suffix = Path(file.filename or "").suffix.lower()

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(status_code=400, detail="Importacao .xlsx requer a dependencia openpyxl no backend.")
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip().lower() for cell in rows[0]]
        result = []
        for values in rows[1:]:
            row = {headers[index]: str(value or "").strip() for index, value in enumerate(values) if index < len(headers)}
            if any(row.values()):
                result.append(row)
        return result

    text_payload = content.decode("utf-8-sig", errors="replace")
    sample = text_payload[:2048]
    delimiter = ";"
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=";,	,").delimiter
    except csv.Error:
        if "," in sample and ";" not in sample:
            delimiter = ","
        elif "\t" in sample:
            delimiter = "\t"
    reader = csv.DictReader(io.StringIO(text_payload), delimiter=delimiter)
    return [
        {str(key or "").strip().lower(): str(value or "").strip() for key, value in row.items()}
        for row in reader
        if any(str(value or "").strip() for value in row.values())
    ]


def preview_company_import_rows(rows: list[dict]) -> dict:
    preview = []
    for index, row in enumerate(rows, start=2):
        errors = []
        warnings = []
        name = row.get("nome_empresa", "").strip()
        status = normalize_status(row.get("status"))
        if not name:
            errors.append("nome_empresa obrigatorio")
        if status not in {ACTIVE_STATUS, INACTIVE_STATUS}:
            errors.append("status invalido")
        preview.append(
            {
                "row": index,
                "data": {
                    "nome_empresa": name,
                    "cnpj": row.get("cnpj", "").strip(),
                    "responsavel": row.get("responsavel", "").strip(),
                    "telefone_1": row.get("telefone_1", "").strip(),
                    "telefone_2": row.get("telefone_2", "").strip(),
                    "status": status,
                },
                "errors": errors,
                "warnings": warnings,
                "valid": not errors,
            }
        )
    return {
        "rows": preview,
        "valid_count": len([row for row in preview if row["valid"]]),
        "error_count": len([row for row in preview if row["errors"]]),
    }


def preview_user_import_rows(rows: list[dict], db) -> dict:
    preview = []
    for index, row in enumerate(rows, start=2):
        errors = []
        warnings = []
        email = row.get("email", "").strip().lower()
        company_id = row.get("id_empresa", "").strip()
        role = normalize_platform_role(row.get("nivel_usuario"))
        if not row.get("nome_usuario", "").strip():
            errors.append("nome_usuario obrigatorio")
        if not EMAIL_PATTERN.match(email):
            errors.append("email invalido")
        company = db.query(Company).filter(Company.id == company_id).first() if company_id else None
        if not company:
            errors.append("id_empresa nao encontrado")
        if role not in PLATFORM_ROLES:
            errors.append("nivel_usuario invalido")
        if not row.get("senha_primaria", "").strip():
            errors.append("senha_primaria obrigatoria")
        existing_user = db.query(User).filter(User.email == email).first() if email else None
        if existing_user:
            warnings.append("email existente: sera criado apenas o vinculo se necessario")
        preview.append(
            {
                "row": index,
                "data": {
                    "nome_usuario": row.get("nome_usuario", "").strip(),
                    "email": email,
                    "senha_primaria": row.get("senha_primaria", "").strip(),
                    "id_empresa": company_id,
                    "telefone": row.get("telefone", "").strip(),
                    "setor": row.get("setor", "").strip() or "Geral",
                    "nivel_usuario": role,
                    "status": normalize_status(row.get("status")),
                },
                "errors": errors,
                "warnings": warnings,
                "valid": not errors,
            }
        )
    return {
        "rows": preview,
        "valid_count": len([row for row in preview if row["valid"]]),
        "error_count": len([row for row in preview if row["errors"]]),
    }


def build_platform_mind_map(db) -> dict:
    companies = db.query(Company).order_by(Company.name.asc()).all()
    return {
        "id": "platform-root",
        "type": "platform",
        "label": "Admin Master",
        "children": [
            {
                "id": company.id,
                "type": "company",
                "label": company.name,
                "status": company.status,
                "children": [
                    {
                        "id": department.id,
                        "type": "department",
                        "label": department.name,
                        "status": department.status,
                        "children": [
                            {
                                "id": str(link.user_id),
                                "link_id": link.id,
                                "type": "user",
                                "label": user.full_name if user else f"Usuario {link.user_id}",
                                "email": user.email if user else None,
                                "role": link.role,
                                "status": link.status,
                            }
                            for link, user in (
                                db.query(CompanyUser, User)
                                .join(User, User.id == CompanyUser.user_id)
                                .filter(CompanyUser.company_id == company.id, CompanyUser.department_id == department.id)
                                .order_by(User.full_name.asc())
                                .all()
                            )
                        ],
                    }
                    for department in (
                        db.query(Department)
                        .filter(Department.company_id == company.id)
                        .order_by(Department.name.asc())
                        .all()
                    )
                ],
            }
            for company in companies
        ],
    }


@app.get("/platform/overview")
async def platform_overview(current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    with SessionLocal() as db:
        companies = db.query(Company).order_by(Company.name.asc()).all()
        users = db.query(User).order_by(User.full_name.asc()).all()
        departments = db.query(Department).order_by(Department.name.asc()).all()
        links = db.query(CompanyUser).order_by(CompanyUser.created_at.desc()).all()
        logs = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(80).all()
        return {
            "stats": {
                "companies": len(companies),
                "active_companies": len([company for company in companies if company.status == ACTIVE_STATUS]),
                "users": len(users),
                "active_users": len([user for user in users if user.is_active]),
                "links": len([link for link in links if link.status == ACTIVE_STATUS]),
                "departments": len(departments),
                "tickets": db.query(Ticket).count(),
                "messages": db.query(Message).count(),
            },
            "companies": [serialize_company(company) for company in companies],
            "users": [serialize_user(user, db) for user in users],
            "departments": [serialize_department(department) for department in departments],
            "company_users": [serialize_company_user(link, db) for link in links],
            "audit_logs": [serialize_audit_log(log) for log in logs],
            "mind_map": build_platform_mind_map(db),
        }


@app.get("/platform/companies")
async def platform_list_companies(current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    with SessionLocal() as db:
        return [serialize_company(company) for company in db.query(Company).order_by(Company.name.asc()).all()]


@app.post("/platform/companies")
async def platform_create_company(payload: dict, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    name = str(payload.get("name") or payload.get("nome_empresa") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nome da empresa e obrigatorio.")
    with SessionLocal() as db:
        cnpj = str(payload.get("cnpj") or "").strip() or None
        if cnpj and db.query(Company).filter(Company.cnpj == cnpj).first():
            raise HTTPException(status_code=400, detail="CNPJ ja cadastrado.")
        company = Company(
            id=str(uuid.uuid4()),
            name=name,
            cnpj=cnpj,
            responsible_name=str(payload.get("responsible_name") or payload.get("responsavel") or "").strip() or None,
            phone_primary=str(payload.get("phone_primary") or payload.get("telefone_1") or "").strip() or None,
            phone_secondary=str(payload.get("phone_secondary") or payload.get("telefone_2") or "").strip() or None,
            status=normalize_status(payload.get("status")),
        )
        db.add(company)
        db.flush()
        for department_name in DEFAULT_DEPARTMENTS:
            get_or_create_department(db, company.id, department_name)
        log_audit(db, current_user.id, "empresa_criada", "company", company.id, company.id, {"name": company.name})
        db.commit()
        db.refresh(company)
        return serialize_company(company)


@app.patch("/platform/companies/{company_id}")
async def platform_update_company(company_id: str, payload: dict, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    with SessionLocal() as db:
        company = db.query(Company).filter(Company.id == company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail="Empresa nao encontrada.")
        for field in ["name", "cnpj", "responsible_name", "phone_primary", "phone_secondary"]:
            if field in payload:
                setattr(company, field, str(payload.get(field) or "").strip() or None)
        if "status" in payload:
            company.status = normalize_status(payload.get("status"))
        company.updated_at = datetime.utcnow()
        log_audit(db, current_user.id, "empresa_editada", "company", company.id, company.id, {"name": company.name})
        db.commit()
        db.refresh(company)
        return serialize_company(company)


@app.get("/platform/departments")
async def platform_list_departments(company_id: Optional[str] = Query(default=None), current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    with SessionLocal() as db:
        query = db.query(Department)
        if company_id:
            query = query.filter(Department.company_id == company_id)
        return [serialize_department(item) for item in query.order_by(Department.name.asc()).all()]


@app.post("/platform/departments")
async def platform_create_department(payload: dict, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    company_id = str(payload.get("company_id") or "").strip()
    name = str(payload.get("name") or "").strip()
    if not company_id or not name:
        raise HTTPException(status_code=400, detail="Empresa e nome do setor sao obrigatorios.")
    with SessionLocal() as db:
        ensure_company_access(db, current_user, company_id)
        department = get_or_create_department(db, company_id, name, str(payload.get("description") or "").strip() or None)
        log_audit(db, current_user.id, "setor_criado", "department", department.id, company_id, {"name": department.name})
        db.commit()
        db.refresh(department)
        return serialize_department(department)


@app.post("/platform/users")
async def platform_create_user(payload: dict, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    email = str(payload.get("email") or "").strip().lower()
    password = str(payload.get("password") or payload.get("senha_primaria") or "").strip()
    company_id = str(payload.get("company_id") or "").strip()
    if not EMAIL_PATTERN.match(email) or not password or not company_id:
        raise HTTPException(status_code=400, detail="Email, senha primaria e empresa sao obrigatorios.")
    role = normalize_platform_role(payload.get("role") or payload.get("nivel_usuario") or "user")
    if role == "master_admin":
        raise HTTPException(status_code=400, detail="Use is_platform_admin para criar Admin Master global.")
    with SessionLocal() as db:
        ensure_company_access(db, current_user, company_id)
        department = get_or_create_department(db, company_id, str(payload.get("department") or payload.get("setor") or "Geral"))
        username = str(payload.get("username") or email.split("@")[0]).strip().lower()
        user = db.query(User).filter(User.email == email).first()
        if not user:
            if db.query(User).filter(User.username == username).first():
                username = f"{username}.{uuid.uuid4().hex[:6]}"
            user = User(
                uuid=str(uuid.uuid4()),
                username=username,
                email=email,
                full_name=str(payload.get("name") or payload.get("full_name") or payload.get("nome_usuario") or email).strip(),
                hashed_password=get_password_hash(password),
                phone=str(payload.get("phone") or payload.get("telefone") or "").strip() or None,
                phone_extension=str(payload.get("phone") or payload.get("telefone") or "").strip() or None,
                must_change_password=True,
                status=normalize_status(payload.get("status")),
                is_active=normalize_status(payload.get("status")) == ACTIVE_STATUS,
                access_level=legacy_access_for_role(role),
                department=department.name,
            )
            db.add(user)
            db.flush()
            log_audit(db, current_user.id, "usuario_criado", "user", user.id, company_id, {"email": user.email})
        link = (
            db.query(CompanyUser)
            .filter(CompanyUser.company_id == company_id, CompanyUser.user_id == user.id)
            .first()
        )
        if not link:
            link = CompanyUser(
                id=str(uuid.uuid4()),
                company_id=company_id,
                user_id=user.id,
                department_id=department.id,
                role=role,
                status=normalize_status(payload.get("status")),
            )
            db.add(link)
            log_audit(db, current_user.id, "vinculo_criado", "company_user", user.id, company_id, {"role": role})
        db.commit()
        db.refresh(user)
        return serialize_user(user, db)


@app.patch("/platform/company-users/{link_id}")
async def platform_update_company_user(link_id: str, payload: dict, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    with SessionLocal() as db:
        link = db.query(CompanyUser).filter(CompanyUser.id == link_id).first()
        if not link:
            raise HTTPException(status_code=404, detail="Vinculo nao encontrado.")
        if "department_id" in payload:
            department = db.query(Department).filter(Department.id == str(payload.get("department_id")), Department.company_id == link.company_id).first()
            if not department:
                raise HTTPException(status_code=400, detail="Setor nao encontrado.")
            link.department_id = department.id
            log_audit(db, current_user.id, "setor_alterado", "company_user", link.id, link.company_id, {"department_id": department.id})
        if "role" in payload:
            link.role = normalize_platform_role(payload.get("role"))
            log_audit(db, current_user.id, "nivel_alterado", "company_user", link.id, link.company_id, {"role": link.role})
        if "status" in payload:
            link.status = normalize_status(payload.get("status"))
        db.commit()
        db.refresh(link)
        return serialize_company_user(link, db)


@app.delete("/platform/company-users/{link_id}")
async def platform_remove_company_user(link_id: str, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    with SessionLocal() as db:
        link = db.query(CompanyUser).filter(CompanyUser.id == link_id).first()
        if not link:
            raise HTTPException(status_code=404, detail="Vinculo nao encontrado.")
        link.status = INACTIVE_STATUS
        log_audit(db, current_user.id, "vinculo_removido", "company_user", link.id, link.company_id, {"user_id": link.user_id})
        db.commit()
        return {"message": "Vinculo removido."}


@app.post("/platform/users/{user_id}/reset-password")
async def platform_reset_user_password(user_id: int, payload: dict, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    new_password = str(payload.get("password") or payload.get("senha_primaria") or "").strip()
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Senha primaria deve ter pelo menos 6 caracteres.")
    with SessionLocal() as db:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuario nao encontrado.")
        user.hashed_password = get_password_hash(new_password)
        user.must_change_password = True
        user.updated_at = datetime.utcnow()
        company_id = payload.get("company_id") or (active_company_memberships(db, user)[0].company_id if active_company_memberships(db, user) else None)
        log_audit(db, current_user.id, "senha_resetada", "user", user.id, company_id, {"email": user.email})
        db.commit()
        return {"message": "Senha resetada.", "user": serialize_user(user, db)}


@app.post("/platform/import/companies/preview")
async def platform_import_companies_preview(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    rows = await read_import_rows(file)
    return preview_company_import_rows(rows)


@app.post("/platform/import/companies/confirm")
async def platform_import_companies_confirm(payload: dict, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    preview = preview_company_import_rows(payload.get("rows") or [])
    if preview["error_count"]:
        return {**preview, "imported": 0, "message": "Corrija os erros antes de importar."}
    imported = 0
    with SessionLocal() as db:
        for row in preview["rows"]:
            data = row["data"]
            company = db.query(Company).filter(Company.cnpj == data["cnpj"]).first() if data["cnpj"] else None
            if not company:
                company = Company(id=str(uuid.uuid4()), cnpj=data["cnpj"] or None)
                db.add(company)
                imported += 1
            company.name = data["nome_empresa"]
            company.responsible_name = data["responsavel"] or None
            company.phone_primary = data["telefone_1"] or None
            company.phone_secondary = data["telefone_2"] or None
            company.status = data["status"]
            company.updated_at = datetime.utcnow()
            for department_name in DEFAULT_DEPARTMENTS:
                get_or_create_department(db, company.id, department_name)
            log_audit(db, current_user.id, "empresa_importada", "company", company.id, company.id, {"name": company.name})
        db.commit()
    return {**preview, "imported": imported}


@app.post("/platform/import/users/preview")
async def platform_import_users_preview(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    rows = await read_import_rows(file)
    with SessionLocal() as db:
        return preview_user_import_rows(rows, db)


@app.post("/platform/import/users/confirm")
async def platform_import_users_confirm(payload: dict, current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    with SessionLocal() as db:
        preview = preview_user_import_rows(payload.get("rows") or [], db)
        if preview["error_count"]:
            return {**preview, "imported": 0, "linked": 0, "message": "Corrija os erros antes de importar."}
        imported = 0
        linked = 0
        for row in preview["rows"]:
            data = row["data"]
            company_id = data["id_empresa"]
            department = get_or_create_department(db, company_id, data["setor"])
            user = db.query(User).filter(User.email == data["email"]).first()
            if not user:
                username = data["email"].split("@")[0].lower()
                if db.query(User).filter(User.username == username).first():
                    username = f"{username}.{uuid.uuid4().hex[:6]}"
                user = User(
                    uuid=str(uuid.uuid4()),
                    username=username,
                    email=data["email"],
                    full_name=data["nome_usuario"],
                    hashed_password=get_password_hash(data["senha_primaria"]),
                    phone=data["telefone"] or None,
                    phone_extension=data["telefone"] or None,
                    must_change_password=True,
                    status=data["status"],
                    is_active=data["status"] == ACTIVE_STATUS,
                    access_level=legacy_access_for_role(data["nivel_usuario"]),
                    department=department.name,
                )
                db.add(user)
                db.flush()
                imported += 1
                log_audit(db, current_user.id, "usuario_importado", "user", user.id, company_id, {"email": user.email})
            else:
                user.must_change_password = True
                user.updated_at = datetime.utcnow()
            link = (
                db.query(CompanyUser)
                .filter(CompanyUser.company_id == company_id, CompanyUser.user_id == user.id)
                .first()
            )
            if not link:
                db.add(
                    CompanyUser(
                        id=str(uuid.uuid4()),
                        company_id=company_id,
                        user_id=user.id,
                        department_id=department.id,
                        role=data["nivel_usuario"],
                        status=data["status"],
                    )
                )
                linked += 1
                log_audit(db, current_user.id, "vinculo_criado", "company_user", user.id, company_id, {"role": data["nivel_usuario"]})
        db.commit()
    return {**preview, "imported": imported, "linked": linked}


@app.get("/admin/overview")
async def admin_overview(current_user: User = Depends(get_current_user)):
    ensure_admin(current_user)
    with SessionLocal() as db:
        users = db.query(User).order_by(User.full_name.asc()).all()
        tickets = db.query(Ticket).order_by(Ticket.created_at.desc()).limit(50).all()
        messages = db.query(Message).order_by(Message.timestamp.desc()).limit(50).all()
        departments = sorted(
            {item.department for item in users if item.department} | {department for department in DEFAULT_DEPARTMENTS}
        )

        return {
            "stats": {
                "users": len(users),
                "active_users": len([item for item in users if item.is_active]),
                "tickets": db.query(Ticket).count(),
                "open_tickets": db.query(Ticket).filter(Ticket.status != "Resolvido").count(),
                "messages": db.query(Message).count(),
                "departments": len(departments),
            },
            "departments": departments,
            "users": [serialize_user(item) for item in users],
            "tickets": [serialize_ticket(ticket, db) for ticket in tickets],
            "messages": [serialize_message(message, db) for message in messages],
        }


@app.get("/coordinator/overview")
async def coordinator_overview(current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user)
        membership = get_company_membership(db, current_user, company_id)
        if not is_company_admin(db, current_user, company_id) and (not membership or membership.role != "coordinator"):
            raise HTTPException(status_code=403, detail="Acesso restrito a coordenadores.")
        department = user_department_name(db, current_user, company_id)
        users_query = (
            db.query(User)
            .join(CompanyUser, CompanyUser.user_id == User.id)
            .filter(CompanyUser.company_id == company_id, CompanyUser.status == ACTIVE_STATUS)
        )
        tickets_query = db.query(Ticket).filter(Ticket.company_id == company_id)
        messages_query = db.query(Message).filter(Message.company_id == company_id)

        if not is_company_admin(db, current_user, company_id):
            membership = get_company_membership(db, current_user, company_id)
            if membership and membership.department_id:
                users_query = users_query.filter(CompanyUser.department_id == membership.department_id)
            tickets_query = tickets_query.filter(Ticket.department == department)
            department_user_ids = [row[0] for row in users_query.with_entities(User.id).all()]
            if department_user_ids:
                messages_query = messages_query.filter(
                    or_(Message.sender_id.in_(department_user_ids), Message.receiver_id.in_(department_user_ids))
                )
            else:
                messages_query = messages_query.filter(Message.sender_id == current_user.id)

        users = users_query.order_by(User.full_name.asc()).all()
        tickets = tickets_query.order_by(Ticket.created_at.desc()).limit(50).all()
        messages = messages_query.order_by(Message.timestamp.desc()).limit(30).all()

        return {
            "department": department,
            "stats": {
                "users": len(users),
                "tickets": len(tickets),
                "open_tickets": len([ticket for ticket in tickets if ticket.status != "Resolvido"]),
                "messages": len(messages),
            },
            "users": [serialize_user(item) for item in users],
            "tickets": [serialize_ticket(ticket, db) for ticket in tickets],
            "messages": [serialize_message(message, db) for message in messages],
        }


@app.get("/groups/")
async def list_groups(company_id: Optional[str] = Query(default=None), current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        scoped_company_id = ensure_company_access(db, current_user, company_id)
        query = db.query(ChatGroup).filter(ChatGroup.company_id == scoped_company_id, ChatGroup.is_active == True)
        if not is_company_admin(db, current_user, scoped_company_id):
            department_name = user_department_name(db, current_user, scoped_company_id)
            if department_name:
                query = query.filter(ChatGroup.department == department_name)
            else:
                query = query.filter(ChatGroup.created_by_id == current_user.id)
        groups = query.order_by(ChatGroup.name.asc()).all()
        return [serialize_group(group, db) for group in groups]


@app.post("/groups/")
async def create_group(payload: dict, current_user: User = Depends(get_current_user)):
    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nome do grupo e obrigatorio.")

    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user, payload.get("company_id"))
        department = str(payload.get("department") or user_department_name(db, current_user, company_id) or "Geral").strip()
        if not is_company_admin(db, current_user, company_id) and not user_can_access_department(db, current_user, company_id, department):
            raise HTTPException(status_code=403, detail="Coordenador so pode criar grupos do proprio setor.")
        get_or_create_department(db, company_id, department)
        group = ChatGroup(
            company_id=company_id,
            name=name,
            description=str(payload.get("description") or "").strip() or None,
            department=department,
            created_by_id=current_user.id,
        )
        db.add(group)
        db.commit()
        db.refresh(group)
        return serialize_group(group, db)


@app.websocket("/messages/ws/{token}")
async def websocket_endpoint(websocket: WebSocket, token: str):
    user_data = None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            await websocket.close(code=1008)
            return

        with SessionLocal() as db:
            user = db.query(User).filter(User.id == int(user_id)).first()
            if not user or user.must_change_password:
                await websocket.close(code=1008)
                return
            company_id = ensure_company_access(db, user)
            user_data = {"id": user.id, "username": user.username, "full_name": user.full_name, "company_id": company_id}

        await manager.connect(websocket, user_data["id"], user_data["company_id"])
        await websocket.send_text(json.dumps({"type": "connection", "message": f"Conectado como {user_data['full_name']}!"}))

        with SessionLocal() as db:
            recent_messages = (
                db.query(Message)
                .filter(Message.company_id == user_data["company_id"])
                .order_by(Message.timestamp.desc())
                .limit(50)
                .all()
            )
            messages_data = [serialize_message(msg, db) for msg in reversed(recent_messages)]
        await websocket.send_text(json.dumps({"type": "message_history", "messages": messages_data}))

        while True:
            data = await websocket.receive_text()
            message_data = json.loads(data)

            if message_data["type"] == "chat_message":
                with SessionLocal() as db:
                    receiver_id = message_data.get("receiver_id")
                    if receiver_id:
                        ensure_user_in_company(db, int(receiver_id), user_data["company_id"])
                    new_message = Message(
                        company_id=user_data["company_id"],
                        content=message_data["content"],
                        sender_id=user_data["id"],
                        receiver_id=receiver_id,
                        message_type=message_data.get("message_type", "text"),
                        file_path=message_data.get("file_path"),
                    )
                    db.add(new_message)
                    db.commit()
                    db.refresh(new_message)
                    message_payload = serialize_message(new_message, db)
                    message_payload["client_id"] = message_data.get("client_id")
                    broadcast_message = {
                        "type": "new_message",
                        "message": message_payload,
                    }

                if broadcast_message["message"]["receiver_id"]:
                    await manager.send_personal_message(json.dumps(broadcast_message), broadcast_message["message"]["receiver_id"])
                    await manager.send_personal_message(json.dumps(broadcast_message), user_data["id"])
                else:
                    await manager.broadcast(json.dumps(broadcast_message), company_id=user_data["company_id"])

            elif message_data["type"] == "typing":
                typing_message = {
                    "type": "typing",
                    "user_id": user_data["id"],
                    "username": user_data["username"],
                    "is_typing": message_data["is_typing"],
                }
                if message_data.get("receiver_id"):
                    await manager.send_personal_message(json.dumps(typing_message), message_data["receiver_id"])
                else:
                    await manager.broadcast(json.dumps(typing_message), exclude_user=user_data["id"], company_id=user_data["company_id"])

            elif message_data["type"] == "ping":
                await websocket.send_text(json.dumps({"type": "pong"}))

    except WebSocketDisconnect:
        pass
    except Exception as exc:
        print(f"Erro na conexao WebSocket: {exc}")
    finally:
        if user_data:
            await manager.broadcast_user_status(user_data["id"], False)
            manager.disconnect(user_data["id"])


@app.post("/files/upload")
async def upload_file(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    file_extension = Path(file.filename or "").suffix.lower()
    if file_extension not in ALLOWED_UPLOAD_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_UPLOAD_EXTENSIONS))
        raise HTTPException(status_code=400, detail=f"Formato nao permitido. Envie apenas: {allowed}.")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 10 MB.")

    upload_dir = Path(os.getenv("UPLOAD_DIR", "uploads"))
    upload_dir.mkdir(parents=True, exist_ok=True)
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = upload_dir / unique_filename
    file_path.write_bytes(content)

    with SessionLocal() as db:
        company_id = ensure_company_access(db, current_user)
        file_record = FileUpload(
            company_id=company_id,
            filename=file.filename or unique_filename,
            file_path=str(file_path),
            file_size=len(content),
            content_type=file.content_type or "application/octet-stream",
            binary_data=content,
            uploaded_by=current_user.id,
        )
        db.add(file_record)
        db.commit()
        db.refresh(file_record)
        return {
            "id": file_record.id,
            "company_id": file_record.company_id,
            "filename": file_record.filename,
            "file_path": file_record.file_path,
            "file_size": file_record.file_size,
            "content_type": file_record.content_type,
        }


@app.get("/files/")
async def list_files(
    company_id: Optional[str] = Query(default=None),
    limit: int = Query(default=50, ge=1, le=100),
    current_user: User = Depends(get_current_user),
):
    with SessionLocal() as db:
        scoped_company_id = ensure_company_access(db, current_user, company_id)
        files = (
            db.query(FileUpload)
            .filter(FileUpload.company_id == scoped_company_id)
            .order_by(FileUpload.upload_date.desc())
            .limit(limit)
            .all()
        )
        return [
            {
                "id": item.id,
                "company_id": item.company_id,
                "filename": item.filename,
                "file_path": item.file_path,
                "file_size": item.file_size,
                "content_type": item.content_type,
                "uploaded_by": item.uploaded_by,
                "upload_date": item.upload_date.isoformat() if item.upload_date else None,
            }
            for item in files
        ]


@app.get("/files/download/{file_id}")
async def download_file(file_id: int, current_user: User = Depends(get_current_user)):
    with SessionLocal() as db:
        file_record = db.query(FileUpload).filter(FileUpload.id == file_id).first()
        if not file_record:
            raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
        ensure_company_access(db, current_user, file_record.company_id)
        file_path = file_record.file_path
        filename = file_record.filename
        content_type = file_record.content_type
        binary_data = bytes(file_record.binary_data or b"") if file_record.binary_data else None

    if not os.path.exists(file_path):
        if not binary_data:
            raise HTTPException(status_code=404, detail="Arquivo nao existe")
        return StreamingResponse(
            iter([binary_data]),
            media_type=content_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(binary_data)),
                "Cache-Control": "no-store",
            },
        )

    return FileResponse(path=file_path, filename=filename, media_type=content_type)


if __name__ == "__main__":
    create_default_users()
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", "8001")))
